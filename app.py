"""
Ozark Finances Flask Application
A web port of the AutoHotkey finance management application
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
import sqlite3
import os
import glob
from datetime import datetime, date, timedelta
import csv
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from pathlib import Path
import logging
import traceback
from config import config
import tempfile
import zipfile
from werkzeug.utils import secure_filename
from audit_tracker import audit_tracker, audit_log, audit_transaction
import math
import statistics

app = Flask(__name__)

# Load configuration
config_name = os.environ.get('FLASK_CONFIG') or 'default'
app.config.from_object(config[config_name])
config[config_name].init_app(app)

# Configure additional upload settings (for templates)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}

# Logging setup
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Create necessary directories
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['GENERATED_FOLDER'], exist_ok=True)

# Add custom Jinja2 filters
def escapejs_filter(value):
    """Escape JavaScript strings to prevent syntax errors"""
    if value is None:
        return ''
    return str(value).replace('\\', '\\\\').replace("'", "\\'").replace('"', '\\"').replace('\n', '\\n').replace('\r', '\\r')

# Register the filter with the Flask app
app.jinja_env.filters['escapejs'] = escapejs_filter

# Alternative registration method to ensure it's available
@app.template_filter('escapejs')
def escapejs_template_filter(value):
    """Alternative registration of escapejs filter"""
    return escapejs_filter(value)

# Custom template filters
@app.template_filter('date_to_html_input')
def date_to_html_input(date_str):
    """Convert DD-MM-YYYY date to YYYY-MM-DD for HTML date input"""
    if not date_str:
        return ''
    try:
        # Parse DD-MM-YYYY format
        date_obj = datetime.strptime(str(date_str), '%d-%m-%Y')
        # Return YYYY-MM-DD format for HTML input
        return date_obj.strftime('%Y-%m-%d')
    except ValueError:
        # If already in YYYY-MM-DD format or invalid, return as-is
        return str(date_str)

@app.template_filter('strptime')
def strptime_filter(date_str, format):
    """Parse datetime string with given format"""
    try:
        return datetime.strptime(str(date_str), format)
    except (ValueError, TypeError):
        return datetime.now()

@app.template_global()
def now():
    """Return current datetime"""
    return datetime.now()

class DatabaseManager:
    def __init__(self, db_path):
        self.db_path = db_path
        
    def get_connection(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn
    
    def execute_query(self, query, params=None):
        """Execute a query and return results"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            return cursor.fetchall()
    
    def execute_update(self, query, params=None, table_name=None, record_id=None):
        """Execute an update/insert/delete query with audit tracking"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # For audit tracking, get old values if this is an UPDATE
            old_values = None
            if query.strip().upper().startswith('UPDATE') and table_name and record_id:
                try:
                    old_cursor = conn.cursor()
                    old_cursor.execute(f"SELECT * FROM {table_name} WHERE {self._get_primary_key_column(table_name)} = ?", (record_id,))
                    old_row = old_cursor.fetchone()
                    if old_row:
                        old_values = dict(old_row)
                except Exception as e:
                    logger.warning(f"Could not fetch old values for audit: {e}")
            
            # Execute the query
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            
            conn.commit()
            rowcount = cursor.rowcount
            
            # Audit tracking
            try:
                action = query.strip().split()[0].upper()
                
                # Determine table name and record ID if not provided
                if not table_name:
                    table_name = self._extract_table_name(query)
                
                if not record_id and action == 'INSERT':
                    # For INSERT, try to get the last inserted row ID or use provided params
                    if params and hasattr(cursor, 'lastrowid') and cursor.lastrowid:
                        record_id = cursor.lastrowid
                    elif params:
                        # Try to extract ID from params (assuming first param is often the ID)
                        record_id = params[0]
                
                # Get new values for INSERT/UPDATE
                new_values = None
                if action in ['INSERT', 'UPDATE'] and table_name and record_id:
                    try:
                        new_cursor = conn.cursor()
                        pk_column = self._get_primary_key_column(table_name)
                        new_cursor.execute(f"SELECT * FROM {table_name} WHERE {pk_column} = ?", (record_id,))
                        new_row = new_cursor.fetchone()
                        if new_row:
                            new_values = dict(new_row)
                    except Exception as e:
                        logger.warning(f"Could not fetch new values for audit: {e}")
                
                # Log the audit action
                if table_name and record_id:
                    from flask import request as flask_request
                    audit_tracker.log_action(
                        action=action,
                        table_name=table_name,
                        record_id=record_id,
                        old_values=old_values,
                        new_values=new_values,
                        user_info={'operation': 'database_update', 'query_type': action},
                        request=flask_request if 'flask_request' in globals() else None
                    )
                    
            except Exception as e:
                logger.warning(f"Audit logging failed: {e}")
            
            return rowcount
    
    def _extract_table_name(self, query):
        """Extract table name from SQL query"""
        try:
            query_upper = query.upper().strip()
            if query_upper.startswith('INSERT INTO'):
                # Extract table name from INSERT INTO table_name
                parts = query_upper.split()
                table_index = parts.index('INTO') + 1
                if table_index < len(parts):
                    return parts[table_index].split('(')[0]  # Remove any column specification
            elif query_upper.startswith('UPDATE'):
                # Extract table name from UPDATE table_name SET
                parts = query_upper.split()
                if len(parts) > 1:
                    return parts[1].split()[0]  # Get table name
            elif query_upper.startswith('DELETE FROM'):
                # Extract table name from DELETE FROM table_name
                parts = query_upper.split()
                from_index = parts.index('FROM') + 1
                if from_index < len(parts):
                    return parts[from_index].split()[0]
            return 'UNKNOWN'
        except:
            return 'UNKNOWN'
    
    def _get_primary_key_column(self, table_name):
        """Get primary key column name for a table"""
        pk_map = {
            'Invoices': 'InvoiceID',
            'Withdraw': 'id',
            'DebtRegister': 'id',
            'KwartaalData': 'id'
        }
        return pk_map.get(table_name, 'id')

db_manager = DatabaseManager(app.config['DATABASE_PATH'])

def format_euro(amount):
    """Format amount as Euro currency"""
    return f"€{amount:,.2f}"

def parse_euro(euro_str):
    """Parse Euro string back to float"""
    return float(euro_str.replace('€', '').replace(',', ''))

def format_payment_id(payment_id):
    """Format payment ID with spaces every 4 digits (e.g., 5256 3108 9150 1210)"""
    if not payment_id:
        return '-'
    
    # Remove any existing spaces and non-digit characters except letters/numbers
    clean_id = str(payment_id).replace(' ', '')
    
    # Add space every 4 characters
    formatted = ' '.join([clean_id[i:i+4] for i in range(0, len(clean_id), 4)])
    return formatted

def allowed_file(filename):
    """Check if uploaded file has allowed extension"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_excel_file(filename):
    """Check if uploaded file is an Excel file"""
    if not filename or '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[1].lower()
    return ext in ['xlsx', 'xls']

def get_next_invoice_number():
    """Generate the next invoice number based on existing invoices"""
    try:
        # Get the highest invoice number from the database
        highest = db_manager.execute_query("""
            SELECT MAX(CAST(InvoiceID AS INTEGER)) as max_id 
            FROM Invoices 
            WHERE InvoiceID GLOB '[0-9]*'
        """)
        
        if highest and highest[0]['max_id']:
            next_number = highest[0]['max_id'] + 1
        else:
            # Start with year + 0001 format
            current_year = str(date.today().year)[2:]  # Last 2 digits of year
            next_number = int(current_year + "0001")
        
        return str(next_number)
    except Exception as e:
        logger.error(f"Error generating invoice number: {e}")
        # Fallback to timestamp-based number
        return str(int(datetime.now().timestamp()))[:-3]

def create_excel_from_template(invoice_data, image_path=None):
    """
    Create an Excel file from your existing template with invoice data and optional image
    Returns the path to the generated file
    (Updated to use your actual Excel template)
    """
    try:
        # Look for template file in common locations
        template_paths = [
            'Template.xlsx',  # Current directory
            'templates/Template.xlsx',  # Templates subdirectory
            'static/Template.xlsx',  # Static files
            os.path.join(os.path.dirname(__file__), 'Template.xlsx'),  # Same directory as app.py
            os.path.join(os.path.dirname(__file__), 'templates', 'Template.xlsx'),
            os.path.join(os.path.dirname(__file__), 'static', 'Template.xlsx'),
        ]
        
        template_path = None
        for path in template_paths:
            if os.path.exists(path):
                template_path = path
                break
        
        if template_path:
            # Load your existing template
            logger.info(f"Loading template from: {template_path}")
            workbook = openpyxl.load_workbook(template_path)
            worksheet = workbook.active
        else:
            # Fallback: create a basic template if not found
            logger.warning("Template.xlsx not found, creating basic template")
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Invoice"
            
            # Basic invoice header
            worksheet['A1'] = "INVOICE"
            worksheet['A1'].font = openpyxl.styles.Font(size=20, bold=True)
        
        # Fill in the key data cells (matching FlaskWebApp approach)
        
        # Invoice number and date (C13, C14)
        worksheet['C13'] = invoice_data.get('invoice_number', '')  # Invoice number at C13
        worksheet['C14'] = invoice_data.get('invoice_date', '')     # Invoice date at C14
        
        # Main amount (F18 - primary amount field like FlaskWebApp)
        if 'amount_excl' in invoice_data:
            try:
                amount = float(invoice_data['amount_excl'])
                worksheet['F18'] = amount
            except (ValueError, TypeError):
                worksheet['F18'] = 0
        
        # Add image if provided (at B21 like FlaskWebApp)
        if image_path and os.path.exists(image_path):
            try:
                img = Image(image_path)
                # Resize image to match FlaskWebApp dimensions
                img.width = int(13.36 * 37.8)  # 13.36 cm width
                img.height = int(10.2 * 37.8)  # 10.2 cm height
                worksheet.add_image(img, 'B21')
                logger.info(f"Added image to cell B21: {image_path}")
            except Exception as e:
                logger.error(f"Error adding image to Excel: {e}")
        
        # Generate filename like FlaskWebApp
        invoice_number = invoice_data.get('invoice_number', 'INV')
        current_date = datetime.now().strftime('%d-%m-%Y')
        default_name = 'TijnBakker'  # FlaskWebApp default
        
        filename = invoice_data.get('filename')
        if not filename:
            filename = f"TringTring_{invoice_number}_{current_date}_{default_name}.xlsx"
        elif not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        output_path = os.path.join(app.config['GENERATED_FOLDER'], filename)
        
        # Save the workbook
        workbook.save(output_path)
        logger.info(f"Invoice saved to: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        return None

@app.route('/')
def index():
    """Main dashboard with comprehensive statistics"""
    try:
        # Get invoice statistics (exclude binned invoices)
        invoice_stats = db_manager.execute_query("""
            SELECT 
                COUNT(*) as total_invoices,
                SUM(Incl) as total_amount,
                AVG(Incl) as avg_amount,
                MAX(Incl) as max_amount,
                MIN(Incl) as min_amount
            FROM Invoices
            WHERE (status IS NULL OR status = 'active')
        """)[0]
        
        # Get recent invoices (last 5, exclude binned)
        recent_invoices = db_manager.execute_query("""
            SELECT InvoiceID, InvoiceDate, Incl 
            FROM Invoices 
            WHERE (status IS NULL OR status = 'active') 
            ORDER BY CAST(InvoiceID AS INTEGER) DESC 
            LIMIT 5
        """)
        
        # Get withdraw statistics
        withdraw_stats = db_manager.execute_query("""
            SELECT 
                COUNT(*) as total_withdraws,
                SUM(Amount) as total_amount,
                AVG(Amount) as avg_amount
            FROM Withdraw
        """)[0]
        
        # Get recent withdraws (last 5)
        recent_withdraws = db_manager.execute_query("""
            SELECT Date, Amount 
            FROM Withdraw 
            ORDER BY Date DESC 
            LIMIT 5
        """)
        
        # Get debt statistics
        debt_stats = db_manager.execute_query("""
            SELECT 
                COUNT(*) as total_debts,
                SUM(Amount) as total_debt_amount,
                SUM(OriginalDebt) as original_debt_amount
            FROM DebtRegister
        """)[0]
        
        # Get all debts for display
        all_debts = db_manager.execute_query("""
            SELECT DebtName, Amount, OriginalDebt 
            FROM DebtRegister 
            ORDER BY Amount DESC
        """)
        
        # Get quarterly data count
        quarterly_info = db_manager.execute_query("""
            SELECT COUNT(*) as total_quarters
            FROM KwartaalData
        """)[0]
        
        # Get quarterly performance data (current year)
        current_year = str(date.today().year)
        quarterly_performance = db_manager.execute_query("""
            SELECT 
                CASE 
                    WHEN CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (1,2,3) THEN 'Q1'
                    WHEN CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (4,5,6) THEN 'Q2'
                    WHEN CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (7,8,9) THEN 'Q3'
                    ELSE 'Q4'
                END as quarter,
                COUNT(*) as invoices_count,
                SUM(Excl) as total_excl,
                SUM(BTW) as total_btw,
                SUM(Incl) as total_incl
            FROM Invoices
            WHERE (status IS NULL OR status = 'active') 
            AND substr(InvoiceDate, 7, 4) = ?
            GROUP BY quarter
            ORDER BY quarter
        """, (current_year,))
        
        # Get VAT summary for current year
        vat_summary = db_manager.execute_query("""
            SELECT 
                COUNT(*) as total_invoices,
                SUM(Excl) as total_excl_amount,
                SUM(BTW) as total_vat_amount,
                SUM(Incl) as total_incl_amount,
                AVG(BTW) as avg_vat_per_invoice
            FROM Invoices
            WHERE (status IS NULL OR status = 'active') 
            AND substr(InvoiceDate, 7, 4) = ?
        """, (current_year,))[0]
        
        # Calculate current quarter info
        current_month = date.today().month
        current_quarter = f"Q{(current_month - 1) // 3 + 1}"
        current_quarter_data = None
        for quarter in quarterly_performance:
            if quarter['quarter'] == current_quarter:
                current_quarter_data = quarter
                break
        
        return render_template('index.html',
                             invoice_stats=invoice_stats,
                             recent_invoices=recent_invoices,
                             withdraw_stats=withdraw_stats,
                             recent_withdraws=recent_withdraws,
                             debt_stats=debt_stats,
                             all_debts=all_debts,
                             quarterly_info=quarterly_info,
                             quarterly_performance=quarterly_performance,
                             current_quarter_data=current_quarter_data,
                             current_quarter=current_quarter,
                             vat_summary=vat_summary,
                             current_year=current_year,
                             format_euro=format_euro)
                             
    except Exception as e:
        print(f"Error in dashboard: {e}")
        # Return basic template without data on error
        return render_template('index.html', 
                             error="Failed to load dashboard data",
                             format_euro=format_euro)

@app.route('/api/dashboard-data')
def api_dashboard_data():
    """API endpoint for real-time dashboard data"""
    try:
        # Get invoice statistics (exclude binned invoices)
        invoice_stats = db_manager.execute_query("""
            SELECT 
                COUNT(*) as total_invoices,
                SUM(Incl) as total_amount,
                AVG(Incl) as avg_amount,
                MAX(Incl) as max_amount,
                MIN(Incl) as min_amount
            FROM Invoices
            WHERE (status IS NULL OR status = 'active')
        """)[0]
        
        # Get recent invoices (last 5, exclude binned)
        recent_invoices = db_manager.execute_query("""
            SELECT InvoiceID, InvoiceDate, Incl 
            FROM Invoices 
            WHERE (status IS NULL OR status = 'active') 
            ORDER BY CAST(InvoiceID AS INTEGER) DESC 
            LIMIT 5
        """)
        
        # Get withdraw statistics
        withdraw_stats = db_manager.execute_query("""
            SELECT 
                COUNT(*) as total_withdraws,
                SUM(Amount) as total_amount,
                AVG(Amount) as avg_amount
            FROM Withdraw
        """)[0]
        
        # Get recent withdraws (last 5)
        recent_withdraws = db_manager.execute_query("""
            SELECT Date, Amount 
            FROM Withdraw 
            ORDER BY Date DESC 
            LIMIT 5
        """)
        
        # Get debt statistics
        debt_stats = db_manager.execute_query("""
            SELECT 
                COUNT(*) as total_debts,
                SUM(Amount) as total_debt_amount,
                SUM(OriginalDebt) as original_debt_amount
            FROM DebtRegister
        """)[0]
        
        # Get all debts for display
        all_debts = db_manager.execute_query("""
            SELECT DebtName, Amount, OriginalDebt 
            FROM DebtRegister 
            ORDER BY Amount DESC
        """)
        
        # Get quarterly data count
        quarterly_info = db_manager.execute_query("""
            SELECT COUNT(*) as total_quarters
            FROM KwartaalData
        """)[0]
        
        # Get quarterly performance data (current year) - API version
        current_year = str(date.today().year)
        quarterly_performance = db_manager.execute_query("""
            SELECT 
                CASE 
                    WHEN CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (1,2,3) THEN 'Q1'
                    WHEN CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (4,5,6) THEN 'Q2'
                    WHEN CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (7,8,9) THEN 'Q3'
                    ELSE 'Q4'
                END as quarter,
                COUNT(*) as invoices_count,
                SUM(Excl) as total_excl,
                SUM(BTW) as total_btw,
                SUM(Incl) as total_incl
            FROM Invoices
            WHERE (status IS NULL OR status = 'active') 
            AND substr(InvoiceDate, 7, 4) = ?
            GROUP BY quarter
            ORDER BY quarter
        """, (current_year,))
        
        # Get VAT summary for current year - API version
        vat_summary = db_manager.execute_query("""
            SELECT 
                COUNT(*) as total_invoices,
                SUM(Excl) as total_excl_amount,
                SUM(BTW) as total_vat_amount,
                SUM(Incl) as total_incl_amount,
                AVG(BTW) as avg_vat_per_invoice
            FROM Invoices
            WHERE (status IS NULL OR status = 'active') 
            AND substr(InvoiceDate, 7, 4) = ?
        """, (current_year,))[0]
        
        # Calculate current quarter info - API version
        current_month = date.today().month
        current_quarter = f"Q{(current_month - 1) // 3 + 1}"
        current_quarter_data = None
        for quarter in quarterly_performance:
            if quarter['quarter'] == current_quarter:
                current_quarter_data = quarter
                break
        
        # Format data for JSON response
        data = {
            'invoice_stats': {
                'total_invoices': invoice_stats['total_invoices'] if invoice_stats else 0,
                'total_amount': float(invoice_stats['total_amount']) if invoice_stats and invoice_stats['total_amount'] else 0.0,
                'avg_amount': float(invoice_stats['avg_amount']) if invoice_stats and invoice_stats['avg_amount'] else 0.0,
                'max_amount': float(invoice_stats['max_amount']) if invoice_stats and invoice_stats['max_amount'] else 0.0
            },
            'withdraw_stats': {
                'total_withdraws': withdraw_stats['total_withdraws'] if withdraw_stats else 0,
                'total_amount': float(withdraw_stats['total_amount']) if withdraw_stats and withdraw_stats['total_amount'] else 0.0,
                'avg_amount': float(withdraw_stats['avg_amount']) if withdraw_stats and withdraw_stats['avg_amount'] else 0.0
            },
            'debt_stats': {
                'total_debts': debt_stats['total_debts'] if debt_stats else 0,
                'total_debt_amount': float(debt_stats['total_debt_amount']) if debt_stats and debt_stats['total_debt_amount'] else 0.0,
                'original_debt_amount': float(debt_stats['original_debt_amount']) if debt_stats and debt_stats['original_debt_amount'] else 0.0
            },
            'quarterly_info': {
                'total_quarters': quarterly_info['total_quarters'] if quarterly_info else 0
            },
            'quarterly_performance': [
                {
                    'quarter': q['quarter'],
                    'invoices_count': q['invoices_count'],
                    'total_excl': float(q['total_excl']) if q['total_excl'] else 0.0,
                    'total_btw': float(q['total_btw']) if q['total_btw'] else 0.0,
                    'total_incl': float(q['total_incl']) if q['total_incl'] else 0.0
                } for q in quarterly_performance
            ] if quarterly_performance else [],
            'current_quarter_data': {
                'quarter': current_quarter_data['quarter'] if current_quarter_data else current_quarter,
                'invoices_count': current_quarter_data['invoices_count'] if current_quarter_data else 0,
                'total_excl': float(current_quarter_data['total_excl']) if current_quarter_data and current_quarter_data['total_excl'] else 0.0,
                'total_btw': float(current_quarter_data['total_btw']) if current_quarter_data and current_quarter_data['total_btw'] else 0.0,
                'total_incl': float(current_quarter_data['total_incl']) if current_quarter_data and current_quarter_data['total_incl'] else 0.0
            },
            'vat_summary': {
                'total_invoices': vat_summary['total_invoices'] if vat_summary else 0,
                'total_excl_amount': float(vat_summary['total_excl_amount']) if vat_summary and vat_summary['total_excl_amount'] else 0.0,
                'total_vat_amount': float(vat_summary['total_vat_amount']) if vat_summary and vat_summary['total_vat_amount'] else 0.0,
                'total_incl_amount': float(vat_summary['total_incl_amount']) if vat_summary and vat_summary['total_incl_amount'] else 0.0,
                'avg_vat_per_invoice': float(vat_summary['avg_vat_per_invoice']) if vat_summary and vat_summary['avg_vat_per_invoice'] else 0.0
            },
            'current_quarter': current_quarter,
            'current_year': current_year,
            'recent_invoices': [
                {
                    'InvoiceID': inv['InvoiceID'],
                    'InvoiceDate': inv['InvoiceDate'],
                    'Incl': float(inv['Incl'])
                } for inv in recent_invoices
            ] if recent_invoices else [],
            'recent_withdraws': [
                {
                    'Date': withdraw['Date'],
                    'Amount': float(withdraw['Amount'])
                } for withdraw in recent_withdraws
            ] if recent_withdraws else [],
            'all_debts': [
                {
                    'DebtName': debt['DebtName'],
                    'Amount': float(debt['Amount']),
                    'OriginalDebt': float(debt['OriginalDebt'])
                } for debt in all_debts
            ] if all_debts else [],
            'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        return jsonify(data)
        
    except Exception as e:
        print(f"Error in API dashboard data: {e}")
        return jsonify({'error': 'Failed to load dashboard data'}), 500

@app.route('/invoices')
def invoices():
    """Invoice management page with comprehensive filtering and sorting"""
    # Get all filter parameters
    invoice_number = request.args.get('invoice_number', '')
    invoice_number_startswith = request.args.get('invoice_number_startswith', '')
    quick_search = request.args.get('quick_search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    exact_date = request.args.get('exact_date', '')
    month_filter = request.args.get('month_filter', '')
    year_filter = request.args.get('year_filter', '')
    quarter_filter = request.args.get('quarter_filter', '')
    weekday_filter = request.args.getlist('weekday_filter')
    date_preset = request.args.get('date_preset', '')
    page = request.args.get('page', 1, type=int)
    per_page = 50  # Items per page
    
    # Get sorting parameters
    sort_by = request.args.get('sort_by', 'invoice_number')  # Default sort by invoice number
    sort_dir = request.args.get('sort_dir', 'desc')  # Default descending (newest first)
    
    try:
        from datetime import datetime, timedelta
        
        # Handle date presets
        if date_preset:
            today = datetime.now()
            if date_preset == 'this_month':
                date_from = today.replace(day=1).strftime('%Y-%m-%d')
                date_to = today.strftime('%Y-%m-%d')
            elif date_preset == 'last_month':
                first_day_this_month = today.replace(day=1)
                last_day_last_month = first_day_this_month - timedelta(days=1)
                first_day_last_month = last_day_last_month.replace(day=1)
                date_from = first_day_last_month.strftime('%Y-%m-%d')
                date_to = last_day_last_month.strftime('%Y-%m-%d')
            elif date_preset == 'this_quarter':
                quarter = (today.month - 1) // 3 + 1
                first_month = (quarter - 1) * 3 + 1
                date_from = today.replace(month=first_month, day=1).strftime('%Y-%m-%d')
                date_to = today.strftime('%Y-%m-%d')
            elif date_preset == 'this_year':
                date_from = today.replace(month=1, day=1).strftime('%Y-%m-%d')
                date_to = today.strftime('%Y-%m-%d')
        
        # Build base query with weekday calculation (FIXED for DD-MM-YYYY format)
        # Only show active invoices (not binned)
        base_query = """
        SELECT InvoiceID, InvoiceDate, Excl, BTW, Incl, payment_status,
               substr(InvoiceDate, 4, 2) as Month,
               substr(InvoiceDate, 7, 4) as Year,
               CASE 
                   WHEN CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (1,2,3) THEN 'Q1'
                   WHEN CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (4,5,6) THEN 'Q2'
                   WHEN CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (7,8,9) THEN 'Q3'
                   ELSE 'Q4'
               END as Quarter,
               CASE strftime('%w', substr(InvoiceDate, 7, 4) || '-' || substr(InvoiceDate, 4, 2) || '-' || substr(InvoiceDate, 1, 2))
                   WHEN '0' THEN 'Sunday'
                   WHEN '1' THEN 'Monday'
                   WHEN '2' THEN 'Tuesday'
                   WHEN '3' THEN 'Wednesday'
                   WHEN '4' THEN 'Thursday'
                   WHEN '5' THEN 'Friday'
                   WHEN '6' THEN 'Saturday'
               END as Weekday
        FROM Invoices
        WHERE (status IS NULL OR status = 'active')
        """
        
        count_query = "SELECT COUNT(*) as total FROM Invoices WHERE (status IS NULL OR status = 'active')"
        params = []
        
        where_conditions = []
        
        # Quick search functionality
        if quick_search:
            # Search in invoice number, amount, or date
            quick_search_condition = """(
                InvoiceID LIKE ? OR 
                CAST(Excl AS TEXT) LIKE ? OR 
                CAST(Incl AS TEXT) LIKE ? OR 
                InvoiceDate LIKE ?
            )"""
            where_conditions.append(quick_search_condition)
            search_param = f"%{quick_search}%"
            params.extend([search_param, search_param, search_param, search_param])
        
        # Invoice number filters
        if invoice_number:
            where_conditions.append("InvoiceID = ?")
            params.append(invoice_number)
        
        if invoice_number_startswith:
            where_conditions.append("InvoiceID LIKE ?")
            params.append(f"{invoice_number_startswith}%")
        
        # Date filters (FIXED for DD-MM-YYYY format)
        if exact_date:
            # Convert YYYY-MM-DD input to DD-MM-YYYY for comparison
            try:
                date_parts = exact_date.split('-')  # YYYY-MM-DD
                formatted_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"  # DD-MM-YYYY
                where_conditions.append("InvoiceDate = ?")
                params.append(formatted_date)
            except:
                pass  # Skip invalid date format
        elif date_from or date_to:
            if date_from:
                # Convert YYYY-MM-DD to DD-MM-YYYY for comparison
                try:
                    date_parts = date_from.split('-')
                    formatted_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
                    # For date comparison with DD-MM-YYYY, we need to convert to comparable format
                    where_conditions.append("date(substr(InvoiceDate, 7, 4) || '-' || substr(InvoiceDate, 4, 2) || '-' || substr(InvoiceDate, 1, 2)) >= date(?)")
                    params.append(date_from)
                except:
                    pass
            if date_to:
                try:
                    date_parts = date_to.split('-')
                    formatted_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
                    where_conditions.append("date(substr(InvoiceDate, 7, 4) || '-' || substr(InvoiceDate, 4, 2) || '-' || substr(InvoiceDate, 1, 2)) <= date(?)")
                    params.append(date_to)
                except:
                    pass
        
        # Time-based filters (FIXED for DD-MM-YYYY format)
        if month_filter:
            where_conditions.append("substr(InvoiceDate, 4, 2) = ?")
            params.append(month_filter.zfill(2))
            
        if year_filter:
            where_conditions.append("substr(InvoiceDate, 7, 4) = ?")
            params.append(year_filter)
            
        if quarter_filter:
            if quarter_filter == 'Q1':
                where_conditions.append("CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (1,2,3)")
            elif quarter_filter == 'Q2':
                where_conditions.append("CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (4,5,6)")
            elif quarter_filter == 'Q3':
                where_conditions.append("CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (7,8,9)")
            elif quarter_filter == 'Q4':
                where_conditions.append("CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (10,11,12)")
        
        # Weekday filter (FIXED for DD-MM-YYYY format)
        if weekday_filter:
            weekday_conditions = []
            for day in weekday_filter:
                day_num = {'Sunday': '0', 'Monday': '1', 'Tuesday': '2', 'Wednesday': '3', 
                          'Thursday': '4', 'Friday': '5', 'Saturday': '6'}.get(day)
                if day_num:
                    # Convert DD-MM-YYYY to YYYY-MM-DD for strftime
                    weekday_conditions.append(f"strftime('%w', substr(InvoiceDate, 7, 4) || '-' || substr(InvoiceDate, 4, 2) || '-' || substr(InvoiceDate, 1, 2)) = '{day_num}'")
            if weekday_conditions:
                where_conditions.append(f"({' OR '.join(weekday_conditions)})")
        
        if where_conditions:
            # base_query already contains WHERE (status IS NULL OR status = 'active')
            # So append additional conditions as AND ...
            and_clause = " AND " + " AND ".join(where_conditions)
            base_query += and_clause
            count_query += and_clause
        
        # Get total count for pagination
        total_result = db_manager.execute_query(count_query, params)
        total_items = total_result[0]['total'] if total_result else 0
        
        # Calculate pagination
        total_pages = (total_items + per_page - 1) // per_page
        offset = (page - 1) * per_page
        has_prev = page > 1
        has_next = page < total_pages
        prev_num = page - 1 if has_prev else None
        next_num = page + 1 if has_next else None
        
        # Build sorting clause
        valid_sort_columns = {
            'invoice_number': 'CAST(InvoiceID AS INTEGER)',
            'date': 'date(substr(InvoiceDate, 7, 4) || \'-\' || substr(InvoiceDate, 4, 2) || \'-\' || substr(InvoiceDate, 1, 2))',
            'amount_excl': 'Excl',
            'amount_btw': 'BTW', 
            'amount_incl': 'Incl'
        }
        
        # Validate sort parameters
        if sort_by not in valid_sort_columns:
            sort_by = 'invoice_number'
        if sort_dir not in ['asc', 'desc']:
            sort_dir = 'desc'
            
        sort_column = valid_sort_columns[sort_by]
        sort_clause = f" ORDER BY {sort_column} {sort_dir.upper()}"
        
        # Build final query with pagination
        query = base_query + sort_clause + " LIMIT ? OFFSET ?"
        params.extend([per_page, offset])
        
        invoices_data = db_manager.execute_query(query, params)
        
        # Calculate totals (for all filtered items, not just current page)
        totals_query = "SELECT SUM(Excl) as total_excl, SUM(BTW) as total_btw, SUM(Incl) as total_incl FROM Invoices WHERE (status IS NULL OR status = 'active')"
        if where_conditions:
            totals_query += " AND " + " AND ".join(where_conditions)
        
        totals_result = db_manager.execute_query(totals_query, params[:-2])  # Remove LIMIT and OFFSET params
        totals = totals_result[0] if totals_result else {'total_excl': 0, 'total_btw': 0, 'total_incl': 0}
        
        # Get bin data for integrated bin card (limit to 10 most recent)
        bin_data = db_manager.execute_query("""
            SELECT InvoiceID, InvoiceDate, Excl, BTW, Incl, deleted_at
            FROM Invoices 
            WHERE status = 'binned'
            ORDER BY deleted_at DESC 
            LIMIT 10
        """)
        
        # Get available years and months for dropdowns (FIXED for DD-MM-YYYY format)
        years_query = "SELECT DISTINCT substr(InvoiceDate,7,4) as year FROM Invoices WHERE (status IS NULL OR status = 'active') ORDER BY year DESC"
        available_years = [row['year'] for row in db_manager.execute_query(years_query)]
        
        months_query = "SELECT DISTINCT substr(InvoiceDate,4,2) as month FROM Invoices WHERE (status IS NULL OR status = 'active') ORDER BY CAST(month AS INTEGER)"
        available_months = [row['month'] for row in db_manager.execute_query(months_query)]
        
        # Create pagination object
        pagination = {
            'page': page,
            'per_page': per_page,
            'total': total_items,
            'total_pages': total_pages,
            'has_prev': has_prev,
            'has_next': has_next,
            'prev_num': prev_num,
            'next_num': next_num
        }
        
        return render_template('invoices.html', 
                             invoices=invoices_data,
                             bin_invoices=bin_data,
                             pagination=pagination,
                             total_excl=totals['total_excl'] or 0,
                             total_btw=totals['total_btw'] or 0,
                             total_incl=totals['total_incl'] or 0,
                             # All filter parameters
                             invoice_number=invoice_number,
                             invoice_number_startswith=invoice_number_startswith,
                             quick_search=quick_search,
                             date_from=date_from,
                             date_to=date_to,
                             exact_date=exact_date,
                             month_filter=month_filter,
                             year_filter=year_filter,
                             quarter_filter=quarter_filter,
                             weekday_filter=weekday_filter,
                             date_preset=date_preset,
                             available_years=available_years,
                             available_months=available_months,
                             # Sorting parameters
                             sort_by=sort_by,
                             sort_dir=sort_dir,
                             format_euro=format_euro)
                             
    except Exception as e:
        print(f"Error in invoices route: {e}")
        return render_template('invoices.html', 
                             invoices=[],
                             bin_invoices=[],
                             pagination=None,
                             total_excl=0,
                             total_btw=0,
                             total_incl=0,
                             # Empty filter parameters for error case
                             invoice_number='',
                             invoice_number_startswith='',
                             quick_search='',
                             date_from='',
                             date_to='',
                             exact_date='',
                             month_filter='',
                             year_filter='',
                             quarter_filter='',
                             weekday_filter=[],
                             date_preset='',
                             available_years=[],
                             available_months=[],
                             # Empty sorting parameters for error case
                             sort_by='invoice_number',
                             sort_dir='desc',
                             format_euro=format_euro,
                             error="Failed to load invoice data")

@app.route('/invoices/move-to-bin/<invoice_id>', methods=['POST'])
def move_invoice_to_bin(invoice_id):
    """Move an invoice to the bin"""
    try:
        # Update invoice status and set deletion timestamp
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        db_manager.execute_update(
            "UPDATE Invoices SET status = 'binned', deleted_at = ? WHERE InvoiceID = ?",
            (current_time, invoice_id),
            table_name='Invoices',
            record_id=invoice_id
        )
        flash(f'Invoice {invoice_id} moved to bin', 'success')
    except Exception as e:
        logger.error(f"Error moving invoice to bin: {e}")
        flash(f'Error moving invoice to bin: {str(e)}', 'error')
    
    return redirect(url_for('invoices'))

@app.route('/invoices/restore/<invoice_id>', methods=['POST'])
def restore_invoice_from_bin(invoice_id):
    """Restore an invoice from the bin"""
    try:
        # Update invoice status back to active and clear deletion timestamp
        db_manager.execute_update(
            "UPDATE Invoices SET status = 'active', deleted_at = NULL WHERE InvoiceID = ?",
            (invoice_id,),
            table_name='Invoices',
            record_id=invoice_id
        )
        flash(f'Invoice {invoice_id} restored from bin', 'success')
    except Exception as e:
        logger.error(f"Error restoring invoice: {e}")
        flash(f'Error restoring invoice: {str(e)}', 'error')
    
    return redirect(url_for('invoices_bin'))

@app.route('/invoices/bin')
def invoices_bin():
    """Invoice bin page showing deleted invoices"""
    try:
        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = 50  # Items per page
        
        # Get total count for pagination
        total_result = db_manager.execute_query(
            "SELECT COUNT(*) as total FROM Invoices WHERE status = 'binned'"
        )
        total_items = total_result[0]['total'] if total_result else 0
        
        # Calculate pagination
        total_pages = (total_items + per_page - 1) // per_page
        offset = (page - 1) * per_page
        has_prev = page > 1
        has_next = page < total_pages
        prev_num = page - 1 if has_prev else None
        next_num = page + 1 if has_next else None
        
        # Get binned invoices with deletion timestamps
        binned_invoices = db_manager.execute_query("""
            SELECT InvoiceID, InvoiceDate, Excl, BTW, Incl, deleted_at
            FROM Invoices 
            WHERE status = 'binned'
            ORDER BY deleted_at DESC 
            LIMIT ? OFFSET ?
        """, (per_page, offset))
        
        # Create pagination object
        pagination = {
            'page': page,
            'per_page': per_page,
            'total': total_items,
            'total_pages': total_pages,
            'has_prev': has_prev,
            'has_next': has_next,
            'prev_num': prev_num,
            'next_num': next_num
        }
        
        return render_template('invoices_bin.html',
                             invoices=binned_invoices,
                             pagination=pagination,
                             format_euro=format_euro)
                             
    except Exception as e:
        logger.error(f"Error loading invoice bin: {e}")
        flash(f'Error loading invoice bin: {str(e)}', 'error')
        return render_template('invoices_bin.html',
                             invoices=[],
                             pagination={
                                 'page': 1,
                                 'per_page': 50,
                                 'total': 0,
                                 'total_pages': 0,
                                 'has_prev': False,
                                 'has_next': False,
                                 'prev_num': None,
                                 'next_num': None
                             },
                             format_euro=format_euro)

@app.route('/invoices/permanent-delete/<invoice_id>', methods=['POST'])
def permanent_delete_invoice(invoice_id):
    """Permanently delete an invoice from the bin"""
    try:
        # Verify invoice is in bin
        invoice = db_manager.execute_query(
            "SELECT * FROM Invoices WHERE InvoiceID = ? AND status = 'binned'",
            (invoice_id,)
        )
        
        if not invoice:
            flash(f'Invoice {invoice_id} not found in bin', 'error')
            return redirect(url_for('invoices_bin'))
        
        # Delete permanently
        db_manager.execute_update(
            "DELETE FROM Invoices WHERE InvoiceID = ? AND status = 'binned'",
            (invoice_id,),
            table_name='Invoices',
            record_id=invoice_id
        )
        
        flash(f'Invoice {invoice_id} permanently deleted', 'success')
        
    except Exception as e:
        logger.error(f"Error permanently deleting invoice {invoice_id}: {e}")
        flash(f'Error deleting invoice: {str(e)}', 'error')
    
    return redirect(url_for('invoices_bin'))

@app.route('/invoices/auto-cleanup', methods=['POST'])
def auto_cleanup_bin():
    """Automatically delete invoices older than 30 days from bin"""
    try:
        # Calculate cutoff date (30 days ago)
        cutoff_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
        
        # Get invoices to be deleted
        invoices_to_delete = db_manager.execute_query(
            "SELECT InvoiceID FROM Invoices WHERE status = 'binned' AND deleted_at < ?",
            (cutoff_date,)
        )
        
        if invoices_to_delete:
            # Delete old invoices
            deleted_count = db_manager.execute_update(
                "DELETE FROM Invoices WHERE status = 'binned' AND deleted_at < ?",
                (cutoff_date,)
            )
            
            flash(f'Auto-cleanup completed: {deleted_count} invoice(s) permanently deleted (older than 30 days)', 'info')
        else:
            flash('No invoices found that are older than 30 days in the bin', 'info')
        
    except Exception as e:
        logger.error(f"Error during auto-cleanup: {e}")
        flash(f'Error during auto-cleanup: {str(e)}', 'error')
    
    return redirect(url_for('invoices_bin'))

@app.route('/invoices/bulk-delete', methods=['POST'])
def bulk_delete_invoices():
    """Permanently delete multiple invoices from the bin"""
    try:
        invoice_ids = request.form.getlist('invoice_ids')
        confirmation = request.form.get('confirmation', '').lower()
        
        if confirmation != 'delete':
            flash('Deletion cancelled - confirmation text did not match', 'warning')
            return redirect(url_for('invoices_bin'))
        
        if not invoice_ids:
            flash('No invoices selected for deletion', 'warning')
            return redirect(url_for('invoices_bin'))
        
        # Delete selected invoices permanently
        placeholders = ','.join(['?' for _ in invoice_ids])
        deleted_count = db_manager.execute_update(
            f"DELETE FROM Invoices WHERE InvoiceID IN ({placeholders}) AND status = 'binned'",
            invoice_ids
        )
        
        if deleted_count > 0:
            flash(f'Permanently deleted {deleted_count} invoice(s)', 'success')
        else:
            flash('No invoices were deleted', 'warning')
            
    except Exception as e:
        logger.error(f"Error bulk deleting invoices: {e}")
        flash(f'Error deleting invoices: {str(e)}', 'error')
    
    return redirect(url_for('invoices_bin'))

@app.route('/invoices/edit/<invoice_id>')
def edit_invoice(invoice_id):
    """Edit invoice page"""
    try:
        # Get invoice data
        invoice_data = db_manager.execute_query(
            "SELECT * FROM Invoices WHERE InvoiceID = ? AND status != 'binned'",
            (invoice_id,)
        )
        
        if not invoice_data:
            flash(f'Invoice {invoice_id} not found', 'error')
            return redirect(url_for('invoices'))
        
        invoice = invoice_data[0]
        
        return render_template('edit_invoice.html',
                             invoice=invoice,
                             format_euro=format_euro)
                             
    except Exception as e:
        logger.error(f"Error loading edit page for invoice {invoice_id}: {e}")
        flash(f'Error loading invoice: {str(e)}', 'error')
        return redirect(url_for('invoices'))

@app.route('/invoices/edit/<invoice_id>', methods=['POST'])
def update_invoice(invoice_id):
    """Update invoice with validation"""
    try:
        # Get form data
        invoice_date = request.form.get('invoice_date', '').strip()
        amount_excl_str = request.form.get('amount_excl', '').strip()
        amount_btw_str = request.form.get('amount_btw', '').strip()
        amount_incl_str = request.form.get('amount_incl', '').strip()
        payment_status = request.form.get('payment_status', 'pending')
        
        # Validation
        errors = []
        
        # Validate date
        if not invoice_date:
            errors.append("Invoice date is required")
        else:
            try:
                # Check if date format is correct (DD-MM-YYYY)
                datetime.strptime(invoice_date, '%Y-%m-%d')  # HTML date input gives YYYY-MM-DD
                # Convert to DD-MM-YYYY format for storage
                date_obj = datetime.strptime(invoice_date, '%Y-%m-%d')
                invoice_date = date_obj.strftime('%d-%m-%Y')
            except ValueError:
                errors.append("Invalid date format")
        
        # Parse amounts with European format support
        def parse_amount(amount_str, field_name):
            if not amount_str:
                errors.append(f"{field_name} is required")
                return None
            
            # Handle European decimal format (comma as decimal separator)
            cleaned_amount = amount_str.replace(',', '.')
            try:
                amount = float(cleaned_amount)
                if amount < 0:
                    errors.append(f"{field_name} cannot be negative")
                    return None
                return round(amount, 2)
            except ValueError:
                errors.append(f"Invalid {field_name.lower()} format")
                return None
        
        amount_excl = parse_amount(amount_excl_str, "Amount excluding BTW")
        amount_btw = parse_amount(amount_btw_str, "BTW amount")
        amount_incl = parse_amount(amount_incl_str, "Amount including BTW")
        
        # Validate BTW calculation (allow for small rounding differences)
        if amount_excl is not None and amount_btw is not None and amount_incl is not None:
            calculated_btw = round(amount_excl * 0.21, 2)
            calculated_incl = round(amount_excl + amount_btw, 2)
            
            # Allow for small rounding differences (±0.01)
            if abs(amount_btw - calculated_btw) > 0.01:
                errors.append(f"BTW amount should be approximately €{calculated_btw:.2f} (21% of €{amount_excl:.2f})")
            
            if abs(amount_incl - calculated_incl) > 0.01:
                errors.append(f"Total amount should be €{calculated_incl:.2f} (excl + BTW)")
        
        # Validate payment status
        valid_statuses = ['pending', 'paid', 'overdue', 'draft']
        if payment_status not in valid_statuses:
            errors.append("Invalid payment status")
        
        if errors:
            # Return errors as JSON for AJAX handling
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'status': 'error', 'errors': errors}), 400
            else:
                for error in errors:
                    flash(error, 'error')
                return redirect(url_for('edit_invoice', invoice_id=invoice_id))
        
        # Update invoice in database
        db_manager.execute_update("""
            UPDATE Invoices 
            SET InvoiceDate = ?, Excl = ?, BTW = ?, Incl = ?, payment_status = ?
            WHERE InvoiceID = ?
        """, (invoice_date, amount_excl, amount_btw, amount_incl, payment_status, invoice_id),
        table_name='Invoices',
        record_id=invoice_id)
        
        flash(f'Invoice {invoice_id} updated successfully!', 'success')
        
        # Return success for AJAX
        if request.headers.get('Content-Type') == 'application/json':
            return jsonify({'status': 'success', 'message': f'Invoice {invoice_id} updated successfully!'})
        
        return redirect(url_for('invoices'))
        
    except Exception as e:
        logger.error(f"Error updating invoice {invoice_id}: {e}")
        error_message = f'Error updating invoice: {str(e)}'
        
        if request.headers.get('Content-Type') == 'application/json':
            return jsonify({'status': 'error', 'message': error_message}), 500
        
        flash(error_message, 'error')
        return redirect(url_for('edit_invoice', invoice_id=invoice_id))

@app.route('/invoices/download/<invoice_id>')
def download_invoice(invoice_id):
    """Download a specific invoice as Excel file"""
    try:
        # Get invoice data
        invoice_data = db_manager.execute_query(
            "SELECT * FROM Invoices WHERE InvoiceID = ? AND status != 'binned'",
            (invoice_id,)
        )
        
        if not invoice_data:
            flash(f'Invoice {invoice_id} not found', 'error')
            return redirect(url_for('invoices'))
        
        invoice = invoice_data[0]
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Invoice {invoice_id}"
        
        # Add invoice data to Excel
        ws['A1'] = 'Invoice ID'
        ws['B1'] = invoice['InvoiceID']
        ws['A2'] = 'Date'
        ws['B2'] = invoice['InvoiceDate']
        ws['A3'] = 'Amount (Excl)'
        ws['B3'] = invoice['Excl']
        ws['A4'] = 'BTW'
        ws['B4'] = invoice['BTW']
        ws['A5'] = 'Amount (Incl)'
        ws['B5'] = invoice['Incl']
        ws['A6'] = 'Payment Status'
        # Handle payment_status safely for sqlite3.Row objects
        try:
            payment_status = invoice['payment_status'] if 'payment_status' in invoice.keys() else 'pending'
        except (KeyError, TypeError):
            payment_status = 'pending'
        ws['B6'] = payment_status
        
        # Style the header
        for cell in ws['A1:A6']:
            cell[0].font = openpyxl.styles.Font(bold=True)
        
        # Save to BytesIO object
        from io import BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create filename - ensure safe characters
        date_str = invoice['InvoiceDate'].replace('/', '-').replace('\\', '-')
        filename = f"Invoice_{invoice_id}_{date_str}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error downloading invoice {invoice_id}: {e}")
        flash(f'Error downloading invoice: {str(e)}', 'error')
        return redirect(url_for('invoices'))

@app.route('/withdraws')
def withdraws():
    """Withdraw management page with sorting functionality"""
    try:
        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)  # Show 50 entries per page
        
        # Get sorting parameters
        sort_by = request.args.get('sort_by', 'date')  # Default sort by date
        sort_dir = request.args.get('sort_dir', 'desc')  # Default descending (newest first)
        
        # Get total count for pagination
        total_count = db_manager.execute_query("SELECT COUNT(*) as count FROM Withdraw")[0]['count']
        
        # Calculate offset
        offset = (page - 1) * per_page
        
        # Build sorting clause
        valid_sort_columns = {
            'date': 'date(substr(Date, 7, 4) || \'-\' || substr(Date, 4, 2) || \'-\' || substr(Date, 1, 2))',
            'amount': 'Amount',
            'description': 'Description'
        }
        
        # Validate sort parameters
        if sort_by not in valid_sort_columns:
            sort_by = 'date'
        if sort_dir not in ['asc', 'desc']:
            sort_dir = 'desc'
            
        sort_column = valid_sort_columns[sort_by]
        sort_clause = f"ORDER BY {sort_column} {sort_dir.upper()}"
        
        # Get paginated withdraws data with sorting
        withdraws_data = db_manager.execute_query(
            f"SELECT Date, Amount, Description FROM Withdraw {sort_clause} LIMIT ? OFFSET ?", 
            (per_page, offset)
        )
        
        # Calculate totals (from all records, not just current page)
        all_withdraws = db_manager.execute_query("SELECT Amount FROM Withdraw")
        total_rows = len(all_withdraws)
        total_amount = 0
        if all_withdraws:
            for withdraw in all_withdraws:
                try:
                    amount = float(withdraw['Amount']) if withdraw['Amount'] is not None else 0
                    total_amount += abs(amount)
                except (ValueError, TypeError):
                    logger.warning(f"Invalid amount value in database: {withdraw['Amount']}")
                    continue
        
        # Get most recent date
        recent_data = db_manager.execute_query("SELECT Date FROM Withdraw ORDER BY Date DESC LIMIT 1")
        recent_date = recent_data[0]['Date'] if recent_data else None
        
        # Calculate pagination info
        total_pages = (total_count + per_page - 1) // per_page
        has_prev = page > 1
        has_next = page < total_pages
        
        return render_template('withdraws.html',
                             withdraws=withdraws_data,
                             total_rows=total_rows,
                             total_amount=total_amount,
                             recent_date=recent_date,
                             format_euro=format_euro,
                             # Sorting parameters
                             sort_by=sort_by,
                             sort_dir=sort_dir,
                             pagination={
                                 'page': page,
                                 'per_page': per_page,
                                 'total': total_count,
                                 'total_pages': total_pages,
                                 'has_prev': has_prev,
                                 'has_next': has_next,
                                 'prev_num': page - 1 if has_prev else None,
                                 'next_num': page + 1 if has_next else None
                             })
    except Exception as e:
        logger.error(f"Error loading withdraws: {e}")
        flash(f'Error loading withdraws: {str(e)}', 'error')
        return render_template('withdraws.html',
                             withdraws=[],
                             total_rows=0,
                             total_amount=0,
                             recent_date=None,
                             format_euro=format_euro,
                             # Empty sorting parameters for error case
                             sort_by='date',
                             sort_dir='desc',
                             pagination={
                                 'page': 1,
                                 'per_page': 50,
                                 'total': 0,
                                 'total_pages': 0,
                                 'has_prev': False,
                                 'has_next': False,
                                 'prev_num': None,
                                 'next_num': None
                             })

@app.route('/withdraws/add', methods=['POST'])
def add_withdraw():
    """Add a new withdraw entry with validation"""
    try:
        date = request.form.get('date', '').strip()
        amount_str = request.form.get('amount', '').strip()
        description = request.form.get('description', '').strip()
        
        # Validation
        errors = []
        
        # Validate date
        if not date:
            errors.append("Date is required")
        else:
            try:
                # Validate date format
                datetime.strptime(date, '%Y-%m-%d')
            except ValueError:
                errors.append("Invalid date format")
        
        # Validate amount
        if not amount_str:
            errors.append("Amount is required")
        else:
            try:
                # Handle European decimal format (comma as decimal separator)
                cleaned_amount = amount_str.replace(',', '.')
                amount = float(cleaned_amount)
                if amount <= 0:
                    errors.append("Amount must be positive")
                # Make amount negative for withdraw
                amount = -abs(amount)
            except ValueError:
                errors.append("Invalid amount format")
                amount = None
        
        # Validate description length
        if description and len(description) > 255:
            errors.append("Description must be 255 characters or less")
        
        if errors:
            for error in errors:
                flash(error, 'error')
            return redirect(url_for('withdraws'))
        
        # Insert into database
        result = db_manager.execute_update(
            "INSERT INTO Withdraw (Date, Amount, Description) VALUES (?, ?, ?)",
            (date, amount, description),
            table_name='Withdraw',
            record_id=None  # Will be auto-generated
        )
        
        flash('Withdraw added successfully', 'success')
    except Exception as e:
        logger.error(f"Error adding withdraw: {e}")
        flash(f'Error adding withdraw: {str(e)}', 'error')
    
    return redirect(url_for('withdraws'))

@app.route('/withdraws/upload', methods=['POST'])
def upload_withdraws():
    """Upload withdraws from CSV file with automatic format detection"""
    if 'file' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('withdraws'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('withdraws'))
    
    if file and file.filename.endswith('.csv'):
        try:
            # Save uploaded file temporarily
            temp_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
            file.save(temp_path)
            
            # Read CSV content
            with open(temp_path, 'r', encoding='utf-8') as f:
                content = f.read()
            lines = content.strip().split('\n')
            
            inserted_count = 0
            skipped_count = 0
            
            # Check if this looks like a Knab export (auto-detect and process)
            is_knab_format = any('Rekeningnummer' in line and 'Transactiedatum' in line and 'Omschrijving' in line for line in lines[:5])
            
            if is_knab_format:
                logger.info("Detected Knab CSV format, processing...")
                
                # Process using Knab CSV processor
                result = process_knab_csv_data(temp_path)
                
                if not result.get('success', False):
                    flash(f'Error processing Knab CSV: {result.get("error", "Unknown error")}', 'error')
                    os.remove(temp_path)
                    return redirect(url_for('withdraws'))
                
                # Read the processed CSV file
                processed_file_path = result['output_path']
                with open(processed_file_path, 'r', encoding='utf-8') as f:
                    processed_content = f.read()
                processed_lines = processed_content.strip().split('\n')
                
                # Skip header line if present
                data_lines = processed_lines[1:] if processed_lines and 'Transactiedatum' in processed_lines[0] else processed_lines
                
                for line_num, line in enumerate(data_lines, 1):
                    if line.strip():
                        # Use CSV reader to properly parse comma-delimited data
                        import csv
                        from io import StringIO
                        try:
                            csv_reader = csv.reader(StringIO(line), delimiter=',')
                            parts = next(csv_reader)
                            
                            if len(parts) >= 3:
                                date = parts[0].strip()
                                amount_str = parts[1].strip()
                                description = parts[2].strip() if len(parts) > 2 else ''
                                
                                # Handle European decimal format
                                amount_str = amount_str.replace(',', '.')
                                amount = float(amount_str)
                                
                                # Basic date format validation
                                if len(date) < 8:
                                    skipped_count += 1
                                    continue
                                
                                db_manager.execute_update(
                                    "INSERT INTO Withdraw (Date, Amount, Description) VALUES (?, ?, ?)",
                                    (date, amount, description)
                                )
                                inserted_count += 1
                            else:
                                skipped_count += 1
                        except (ValueError, StopIteration) as e:
                            logger.warning(f"Skipping processed line {line_num}: {e}")
                            skipped_count += 1
                            continue
                
                # Clean up processed file
                if os.path.exists(processed_file_path):
                    os.remove(processed_file_path)
                    
                flash(f'Successfully processed Knab CSV and imported {inserted_count} withdraw entries' + 
                      (f' (skipped {skipped_count} invalid entries)' if skipped_count > 0 else ''), 'success')
            
            else:
                logger.info("Processing as simple CSV format...")
                
                # Process as simple CSV format (date,amount,description)
                for line_num, line in enumerate(lines, 1):
                    if line.strip():
                        # Try comma first, then semicolon
                        parts = line.split(',')
                        if len(parts) < 2:
                            parts = line.split(';')
                        
                        if len(parts) >= 2:
                            try:
                                date = parts[0].strip().strip('"')
                                amount_str = parts[1].strip().strip('"')
                                description = parts[2].strip().strip('"') if len(parts) > 2 else ''
                                
                                # Handle European decimal format (comma as decimal separator)
                                amount_str = amount_str.replace(',', '.')
                                amount = float(amount_str)
                                
                                # Basic date format validation
                                if len(date) < 8:  # Minimum date length
                                    skipped_count += 1
                                    continue
                                
                                db_manager.execute_update(
                                    "INSERT INTO Withdraw (Date, Amount, Description) VALUES (?, ?, ?)",
                                    (date, amount, description)
                                )
                                inserted_count += 1
                            except ValueError as e:
                                logger.warning(f"Skipping line {line_num}: {e}")
                                skipped_count += 1
                                continue
                        else:
                            skipped_count += 1
                
                if inserted_count > 0:
                    flash(f'Successfully imported {inserted_count} withdraw entries' + 
                          (f' (skipped {skipped_count} invalid lines)' if skipped_count > 0 else ''), 'success')
                else:
                    flash('No valid withdraw entries found. Expected format: date,amount,description (one per line)', 'error')
            
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)
                
        except Exception as e:
            logger.error(f"Error uploading withdraws: {e}")
            flash(f'Error processing file: {str(e)}', 'error')
            # Clean up on error
            if 'temp_path' in locals() and os.path.exists(temp_path):
                os.remove(temp_path)
    else:
        flash('Please upload a CSV file', 'error')
    
    return redirect(url_for('withdraws'))

@app.route('/withdraws/delete/<int:row_id>', methods=['POST'])
def delete_withdraw(row_id):
    """Delete a withdraw entry"""
    date = request.form.get('date')
    amount = request.form.get('amount')
    description = request.form.get('description', '')
    
    # Use description in WHERE clause for more precise deletion
    db_manager.execute_update(
        "DELETE FROM Withdraw WHERE Date = ? AND Amount = ? AND (Description = ? OR (Description IS NULL AND ? = ''))",
        (date, amount, description, description),
        table_name='Withdraw',
        record_id=row_id
    )
    
    flash('Withdraw deleted successfully', 'success')
    return redirect(url_for('withdraws'))

@app.route('/important-info')
def important_info():
    """Important quarterly information page"""
    try:
        # Get BTW quarterly payments
        btw_payments = db_manager.execute_query("""
            SELECT id, timeframe, quarter_months, latest_payment_date, 
                   payment_id, cost, actual_payment_date, status, notes
            FROM btw_quarterly_payments 
            ORDER BY latest_payment_date ASC
        """)
        
        # Calculate auto-costs for each quarter from invoice data
        current_year = datetime.now().year
        quarters = [
            (f"Q1 {current_year}", f"{current_year}-01-01", f"{current_year}-03-31"),
            (f"Q2 {current_year}", f"{current_year}-04-01", f"{current_year}-06-30"),
            (f"Q3 {current_year}", f"{current_year}-07-01", f"{current_year}-09-30"),
            (f"Q4 {current_year}", f"{current_year}-10-01", f"{current_year}-12-31"),
        ]
        
        quarter_calculations = {}
        for timeframe, start_date, end_date in quarters:
            btw_total = db_manager.execute_query("""
                SELECT COALESCE(SUM(BTW), 0) as total_btw 
                FROM Invoices 
                WHERE InvoiceDate BETWEEN ? AND ? 
                AND status = 'active'
            """, (start_date, end_date))
            
            quarter_calculations[timeframe] = btw_total[0]['total_btw'] if btw_total else 0.0
        
        return render_template('important_info.html',
                             btw_payments=btw_payments,
                             quarter_calculations=quarter_calculations,
                             format_euro=format_euro,
                             format_payment_id=format_payment_id,
                             today=datetime.now().strftime('%Y-%m-%d'))
    except Exception as e:
        logger.error(f"Error loading important info: {e}")
        flash(f'Error loading important info: {str(e)}', 'error')
        return render_template('important_info.html',
                             btw_payments=[],
                             quarter_calculations={},
                             format_euro=format_euro,
                             format_payment_id=format_payment_id,
                             today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/test-dialogs')
def test_dialogs():
    """Test page for custom dialogs"""
    return render_template('test_dialogs.html')

@app.route('/btw/update/<int:payment_id>', methods=['POST'])
def update_btw_payment(payment_id):
    """Update BTW quarterly payment details"""
    try:
        data = request.get_json()
        
        # Update the BTW payment record
        db_manager.execute_update("""
            UPDATE btw_quarterly_payments 
            SET payment_id = ?, cost = ?, actual_payment_date = ?, 
                status = ?, notes = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (
            data.get('payment_id', ''),
            float(data.get('cost', 0)),
            data.get('actual_payment_date', ''),
            data.get('status', 'pending'),
            data.get('notes', ''),
            payment_id
        ))
        
        return jsonify({'status': 'success', 'message': 'BTW payment updated successfully'})
        
    except Exception as e:
        logger.error(f"Error updating BTW payment: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/btw/mark-paid/<int:payment_id>', methods=['POST'])
def mark_btw_paid(payment_id):
    """Mark BTW payment as paid with current date"""
    try:
        current_date = datetime.now().strftime('%Y-%m-%d')
        
        db_manager.execute_update("""
            UPDATE btw_quarterly_payments 
            SET status = 'paid', actual_payment_date = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (current_date, payment_id))
        
        flash('BTW payment marked as paid!', 'success')
        
    except Exception as e:
        logger.error(f"Error marking BTW payment as paid: {e}")
        flash(f'Error updating payment: {str(e)}', 'error')
    
    return redirect(url_for('important_info'))

@app.route('/btw/reverse-payment/<int:payment_id>', methods=['POST'])
def reverse_btw_payment(payment_id):
    """Reverse a BTW payment (mark as unpaid)"""
    try:
        db_manager.execute_update("""
            UPDATE btw_quarterly_payments 
            SET status = 'pending', actual_payment_date = NULL, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (payment_id,))
        
        flash('BTW payment reversed - marked as unpaid', 'warning')
        
    except Exception as e:
        logger.error(f"Error reversing BTW payment: {e}")
        flash(f'Error reversing payment: {str(e)}', 'error')
    
    return redirect(url_for('important_info'))

@app.route('/btw/auto-calculate/<int:payment_id>', methods=['POST'])
def auto_calculate_btw(payment_id):
    """Auto-calculate BTW cost from invoice data"""
    try:
        # Get the timeframe for this payment
        payment_data = db_manager.execute_query(
            "SELECT timeframe FROM btw_quarterly_payments WHERE id = ?", 
            (payment_id,)
        )
        
        if not payment_data:
            flash('Payment not found', 'error')
            return redirect(url_for('important_info'))
        
        timeframe = payment_data[0]['timeframe']
        
        # Extract year and quarter from timeframe (e.g., "Q1 2025")
        parts = timeframe.split()
        if len(parts) != 2:
            flash('Invalid timeframe format', 'error')
            return redirect(url_for('important_info'))
        
        quarter = parts[0]
        year = int(parts[1])
        
        # Define quarter date ranges
        quarter_ranges = {
            'Q1': (f"{year}-01-01", f"{year}-03-31"),
            'Q2': (f"{year}-04-01", f"{year}-06-30"),
            'Q3': (f"{year}-07-01", f"{year}-09-30"),
            'Q4': (f"{year}-10-01", f"{year}-12-31"),
        }
        
        if quarter not in quarter_ranges:
            flash('Invalid quarter', 'error')
            return redirect(url_for('important_info'))
        
        start_date, end_date = quarter_ranges[quarter]
        
        # Calculate total BTW for the quarter from paid invoices
        # Use quarter and year filtering that matches the existing format
        btw_total = db_manager.execute_query("""
            SELECT COALESCE(SUM(BTW), 0) as total_btw 
            FROM Invoices 
            WHERE substr(InvoiceDate, 7, 4) = ? 
            AND CASE 
                WHEN CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (1,2,3) THEN 'Q1'
                WHEN CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (4,5,6) THEN 'Q2'
                WHEN CAST(substr(InvoiceDate, 4, 2) AS INTEGER) IN (7,8,9) THEN 'Q3'
                ELSE 'Q4'
            END = ?
            AND (status IS NULL OR status = 'active')
            AND (payment_status = 'paid')
        """, (str(year), quarter))
        
        calculated_amount = btw_total[0]['total_btw'] if btw_total else 0.0
        
        # Update the payment with calculated amount
        db_manager.execute_update("""
            UPDATE btw_quarterly_payments 
            SET cost = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (calculated_amount, payment_id))
        
        flash(f'BTW amount auto-calculated: {format_euro(calculated_amount)}', 'info')
        
    except Exception as e:
        logger.error(f"Error auto-calculating BTW: {e}")
        flash(f'Error calculating BTW: {str(e)}', 'error')
    
    return redirect(url_for('important_info'))

# Invoice Payment Status Routes
@app.route('/invoices/mark-paid/<invoice_id>', methods=['POST'])
def mark_invoice_paid(invoice_id):
    """Mark an invoice as paid"""
    try:
        db_manager.execute_update("""
            UPDATE Invoices 
            SET payment_status = 'paid'
            WHERE InvoiceID = ?
        """, (invoice_id,))
        
        flash(f'Invoice {invoice_id} marked as paid!', 'success')
        
    except Exception as e:
        logger.error(f"Error marking invoice as paid: {e}")
        flash(f'Error updating invoice: {str(e)}', 'error')
    
    return redirect(url_for('invoices'))

# Bulk Invoice Operations
@app.route('/invoices/bulk/mark-paid', methods=['POST'])
def bulk_mark_invoices_paid():
    """Mark multiple invoices as paid"""
    try:
        invoice_ids = request.form.getlist('invoice_ids')
        if not invoice_ids:
            flash('No invoices selected', 'warning')
            return redirect(url_for('invoices'))
        
        for invoice_id in invoice_ids:
            db_manager.execute_update("""
                UPDATE Invoices 
                SET payment_status = 'paid'
                WHERE InvoiceID = ?
            """, (invoice_id,))
        
        flash(f'{len(invoice_ids)} invoice(s) marked as paid!', 'success')
        
    except Exception as e:
        logger.error(f"Error marking invoices as paid: {e}")
        flash(f'Error updating invoices: {str(e)}', 'error')
    
    return redirect(url_for('invoices'))

@app.route('/invoices/bulk/move-to-bin', methods=['POST'])
def bulk_move_invoices_to_bin():
    """Move multiple invoices to bin"""
    try:
        invoice_ids = request.form.getlist('invoice_ids')
        if not invoice_ids:
            flash('No invoices selected', 'warning')
            return redirect(url_for('invoices'))
        
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for invoice_id in invoice_ids:
            db_manager.execute_update("""
                UPDATE Invoices 
                SET status = 'binned', deleted_at = ?
                WHERE InvoiceID = ?
            """, (current_time, invoice_id))
        
        flash(f'{len(invoice_ids)} invoice(s) moved to bin!', 'success')
        
    except Exception as e:
        logger.error(f"Error moving invoices to bin: {e}")
        flash(f'Error moving invoices: {str(e)}', 'error')
    
    return redirect(url_for('invoices'))

@app.route('/api/show-message', methods=['POST'])
def api_show_message():
    """API endpoint to show client-side messages"""
    try:
        data = request.get_json()
        message = data.get('message', 'No message provided')
        msg_type = data.get('type', 'info')  # success, error, warning, info, primary
        duration = data.get('duration', 5000)
        method = data.get('method', 'toast')  # toast, modal, alert
        
        return jsonify({
            'status': 'success',
            'message': message,
            'type': msg_type,
            'duration': duration,
            'method': method
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/bin-count')
def api_bin_count():
    """API endpoint to get the current bin count"""
    try:
        result = db_manager.execute_query(
            "SELECT COUNT(*) as count FROM Invoices WHERE status = 'binned'"
        )
        count = result[0]['count'] if result else 0
        return jsonify({'count': count})
    except Exception as e:
        logger.error(f"Error getting bin count: {e}")
        return jsonify({'count': 0})

@app.route('/demo/messages')
def message_demo():
    """Demo page for testing different message types"""
    return render_template('message_demo.html')

@app.route('/demo/invoices-layout')
def demo_invoices_layout():
    """Demo page showing invoices page with integrated bin card"""
    try:
        # Get sample invoice data (limit to 20 for demo)
        invoices_data = db_manager.execute_query("""
            SELECT InvoiceID, InvoiceDate, Excl, BTW, Incl 
            FROM Invoices 
            WHERE (status IS NULL OR status = 'active')
            ORDER BY CAST(InvoiceID AS INTEGER) DESC 
            LIMIT 20
        """)
        
        # Get sample bin data (limit to 10 for demo)
        bin_data = db_manager.execute_query("""
            SELECT InvoiceID, InvoiceDate, Excl, BTW, Incl, deleted_at
            FROM Invoices 
            WHERE status = 'binned'
            ORDER BY deleted_at DESC 
            LIMIT 10
        """)
        
        # Calculate totals for active invoices
        totals_result = db_manager.execute_query("""
            SELECT SUM(Excl) as total_excl, SUM(BTW) as total_btw, SUM(Incl) as total_incl 
            FROM Invoices 
            WHERE (status IS NULL OR status = 'active')
        """)
        
        totals = totals_result[0] if totals_result else {'total_excl': 0, 'total_btw': 0, 'total_incl': 0}
        
        # Create mock pagination object
        pagination = {
            'page': 1,
            'total': len(invoices_data),
            'total_pages': 1,
            'has_prev': False,
            'has_next': False,
            'per_page': 20
        }
        
        return render_template('demo_invoices_layout.html',
                             invoices=invoices_data,
                             bin_invoices=bin_data,
                             pagination=pagination,
                             total_excl=totals['total_excl'] or 0,
                             total_btw=totals['total_btw'] or 0,
                             total_incl=totals['total_incl'] or 0,
                             format_euro=format_euro)
                             
    except Exception as e:
        logger.error(f"Error loading demo invoices layout: {e}")
        return render_template('demo_invoices_layout.html',
                             invoices=[],
                             bin_invoices=[],
                             pagination={'page': 1, 'total': 0, 'total_pages': 0, 'has_prev': False, 'has_next': False, 'per_page': 20},
                             total_excl=0,
                             total_btw=0,
                             total_incl=0,
                             format_euro=format_euro,
                             error="Failed to load demo data")

@app.route('/demo/flash/<msg_type>')
def demo_flash_message(msg_type):
    """Demo flash messages of different types"""
    messages = {
        'success': 'This is a success message! Everything worked perfectly.',
        'error': 'This is an error message! Something went wrong.',
        'warning': 'This is a warning message! Please pay attention.',
        'info': 'This is an info message! Here is some useful information.',
        'primary': 'This is a primary message! This is important.',
        'secondary': 'This is a secondary message! Additional information.'
    }
    
    message = messages.get(msg_type, 'Unknown message type')
    flash(message, msg_type)
    return redirect(url_for('message_demo'))

@app.route('/debt')
def debt():
    """Debt management page with enhanced features"""
    try:
        # Get filter and sort parameters
        category_filter = request.args.get('category', '')
        sort_by = request.args.get('sort_by', 'amount_desc')
        min_amount = request.args.get('min_amount', '', type=str)
        max_amount = request.args.get('max_amount', '', type=str)
        
        # Build base query with all available columns (handle missing columns gracefully)
        base_query = """
            SELECT DebtName, Amount, OriginalDebt, 
                   COALESCE(Category, '') as Category,
                   COALESCE(DueDate, '') as DueDate,
                   COALESCE(MinimumPayment, 0) as MinimumPayment,
                   COALESCE(InterestRate, 0) as InterestRate,
                   COALESCE(Notes, '') as Notes,
                   COALESCE(AddedDate, '') as AddedDate
            FROM DebtRegister
        """
        
        conditions = []
        params = []
        
        # Apply filters
        if category_filter:
            conditions.append("(Category = ? OR (Category IS NULL AND ? = ''))")
            params.extend([category_filter, category_filter])
        
        if min_amount:
            try:
                min_val = float(min_amount)
                conditions.append("Amount >= ?")
                params.append(min_val)
            except ValueError:
                pass
                
        if max_amount:
            try:
                max_val = float(max_amount)
                conditions.append("Amount <= ?")
                params.append(max_val)
            except ValueError:
                pass
        
        if conditions:
            base_query += " WHERE " + " AND ".join(conditions)
        
        # Add sorting
        sort_options = {
            'amount_desc': 'Amount DESC',
            'amount_asc': 'Amount ASC',
            'name_asc': 'DebtName ASC',
            'name_desc': 'DebtName DESC',
            'due_date': 'DueDate ASC',
            'interest_rate': 'InterestRate DESC'
        }
        
        if sort_by in sort_options:
            base_query += f" ORDER BY {sort_options[sort_by]}"
        else:
            base_query += " ORDER BY Amount DESC"  # Default sort
        
        # Execute query
        debt_data = db_manager.execute_query(base_query, params)
        
        # Calculate statistics
        statistics = None
        if debt_data:
            total_current = sum(debt['Amount'] for debt in debt_data)
            total_original = sum(debt['OriginalDebt'] for debt in debt_data)
            total_paid = total_original - total_current
            progress_percent = (total_paid / total_original * 100) if total_original > 0 else 0
            min_monthly = sum(debt['MinimumPayment'] or 0 for debt in debt_data)
            
            statistics = {
                'debt_count': len(debt_data),
                'total_current': total_current,
                'total_original': total_original,
                'total_paid': total_paid,
                'progress_percent': progress_percent,
                'min_monthly': min_monthly
            }
        
        # Get unique categories for filter dropdown
        try:
            categories_query = db_manager.execute_query(
                "SELECT DISTINCT Category FROM DebtRegister WHERE Category IS NOT NULL AND Category != '' ORDER BY Category"
            )
            categories = [cat['Category'] for cat in categories_query] if categories_query else []
        except:
            categories = []
        
        return render_template('debt.html',
                             debt_data=debt_data,
                             statistics=statistics,
                             categories=categories,
                             format_euro=format_euro)
                             
    except Exception as e:
        logger.error(f"Error loading debt page: {e}")
        flash(f'Error loading debt page: {str(e)}', 'error')
        return render_template('debt.html',
                             debt_data=[],
                             statistics=None,
                             categories=[],
                             format_euro=format_euro)

@app.route('/debt/add', methods=['POST'])
def add_debt():
    """Add a new debt with enhanced validation and features"""
    try:
        debt_name = request.form.get('debt_name', '').strip()
        amount_str = request.form.get('amount', '').strip()
        category = request.form.get('category', '').strip()
        due_date = request.form.get('due_date', '').strip()
        min_payment_str = request.form.get('minimum_payment', '').strip()
        interest_rate_str = request.form.get('interest_rate', '').strip()
        notes = request.form.get('notes', '').strip()
        
        # Validation
        errors = []
        
        # Validate debt name
        if not debt_name:
            errors.append("Debt name is required")
        elif len(debt_name) > 21:
            errors.append("Debt name must be 21 characters or less")
        elif not debt_name.replace(' ', '').isalnum():
            errors.append("Debt name can only contain letters, numbers, and spaces")
        
        # Validate amount
        if not amount_str:
            errors.append("Amount is required")
        else:
            try:
                # Handle European decimal format (comma as decimal separator)
                cleaned_amount = amount_str.replace(',', '.')
                amount = float(cleaned_amount)
                if amount <= 0:
                    errors.append("Amount must be positive")
            except ValueError:
                errors.append("Invalid amount format")
                amount = None
        
        # Validate minimum payment (optional)
        min_payment = 0
        if min_payment_str:
            try:
                min_payment = float(min_payment_str.replace(',', '.'))
                if min_payment < 0:
                    errors.append("Minimum payment cannot be negative")
            except ValueError:
                errors.append("Invalid minimum payment format")
        
        # Validate interest rate (optional)
        interest_rate = 0
        if interest_rate_str:
            try:
                interest_rate = float(interest_rate_str.replace(',', '.'))
                if interest_rate < 0 or interest_rate > 100:
                    errors.append("Interest rate must be between 0 and 100")
            except ValueError:
                errors.append("Invalid interest rate format")
        
        # Validate due date (optional)
        if due_date:
            try:
                datetime.strptime(due_date, '%Y-%m-%d')
            except ValueError:
                errors.append("Invalid due date format")
        
        # Validate notes length
        if notes and len(notes) > 500:
            errors.append("Notes must be 500 characters or less")
        
        # Check if debt name already exists
        if debt_name:
            existing = db_manager.execute_query(
                "SELECT COUNT(*) as count FROM DebtRegister WHERE DebtName = ?", 
                (debt_name,)
            )
            if existing and existing[0]['count'] > 0:
                errors.append("A debt with this name already exists")
        
        if errors:
            for error in errors:
                flash(error, 'error')
            return redirect(url_for('debt'))
        
        # Create debt database file with enhanced schema
        debt_db_path = app.config['DATA_DIR'] / f"{debt_name}.sqlite"
        
        with sqlite3.connect(debt_db_path) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS DebtSource (
                    Amount REAL NOT NULL,
                    CreatedDate TEXT NOT NULL,
                    UNIX INTEGER NOT NULL,
                    PaymentMethod TEXT,
                    Notes TEXT
                )
            """)
        
        # Add to main register with new fields (use current date for AddedDate)
        unix_timestamp = int(datetime.now().timestamp())
        current_date = datetime.now().strftime('%Y-%m-%d')
        
        db_manager.execute_update("""
            INSERT INTO DebtRegister 
            (DebtName, Amount, UnixStamp, OriginalDebt, Category, DueDate, MinimumPayment, InterestRate, Notes, AddedDate) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (debt_name, amount, unix_timestamp, amount, 
              category if category else None,
              due_date if due_date else None, 
              min_payment if min_payment > 0 else None, 
              interest_rate if interest_rate > 0 else None, 
              notes if notes else None,
              current_date),
            table_name='DebtRegister',
            record_id=debt_name
        )
        
        flash('Debt added successfully', 'success')
    except Exception as e:
        logger.error(f"Error adding debt: {e}")
        flash(f'Error adding debt: {str(e)}', 'error')
    
    return redirect(url_for('debt'))

@app.route('/debt/update/<debt_name>', methods=['POST'])
def update_debt(debt_name):
    """Update debt amount (payment) with enhanced validation"""
    try:
        payment_amount_str = request.form.get('payment_amount', '').strip()
        payment_method = request.form.get('payment_method', '').strip()
        payment_notes = request.form.get('payment_notes', '').strip()
        
        # Validation
        errors = []
        
        if not payment_amount_str:
            errors.append("Payment amount is required")
            payment_amount = 0
        else:
            try:
                payment_amount = float(payment_amount_str.replace(',', '.'))
                if payment_amount <= 0:
                    errors.append("Payment amount must be positive")
            except ValueError:
                errors.append("Invalid payment amount format")
                payment_amount = 0
        
        # Validate payment notes length
        if payment_notes and len(payment_notes) > 255:
            errors.append("Payment notes must be 255 characters or less")
        
        if errors:
            for error in errors:
                flash(error, 'error')
            return redirect(url_for('debt'))
        
        # Get current debt amount
        current_debt = db_manager.execute_query(
            "SELECT Amount FROM DebtRegister WHERE DebtName = ?", (debt_name,)
        )
        
        if current_debt:
            current_amount = current_debt[0]['Amount']
            new_amount = current_amount - payment_amount
            
            # Handle overpayment
            if new_amount < 0:
                overpayment = abs(new_amount)
                new_amount = 0
                flash(f'Payment recorded. Overpayment of {format_euro(overpayment)} - debt is now fully paid!', 'warning')
            
            # Update main register
            db_manager.execute_update(
                "UPDATE DebtRegister SET Amount = ? WHERE DebtName = ?",
                (new_amount, debt_name),
                table_name='DebtRegister',
                record_id=debt_name
            )
            
            # Add payment record to debt database with enhanced tracking
            debt_db_path = app.config['DATA_DIR'] / f"{debt_name}.sqlite"
            if debt_db_path.exists():
                with sqlite3.connect(debt_db_path) as conn:
                    conn.execute("""
                        INSERT INTO DebtSource (Amount, CreatedDate, UNIX, PaymentMethod, Notes) 
                        VALUES (?, ?, ?, ?, ?)
                    """, (
                        payment_amount, 
                        datetime.now().strftime('%d/%m/%Y'), 
                        int(datetime.now().timestamp()),
                        payment_method,
                        payment_notes
                    ))
            
            if new_amount == 0:
                flash(f'Congratulations! Debt "{debt_name}" is now fully paid!', 'success')
            else:
                flash(f'Payment of {format_euro(payment_amount)} recorded. Remaining debt: {format_euro(new_amount)}', 'success')
        else:
            flash('Debt not found', 'error')
    
    except Exception as e:
        logger.error(f"Error updating debt: {e}")
        flash(f'Error updating debt: {str(e)}', 'error')
    
    return redirect(url_for('debt'))

@app.route('/debt/edit/<debt_name>')
def edit_debt(debt_name):
    """Edit debt page"""
    try:
        # Get current debt data with safe column access
        debt_data = db_manager.execute_query("""
            SELECT DebtName, Amount, OriginalDebt,
                   COALESCE(Category, '') as Category,
                   COALESCE(DueDate, '') as DueDate,
                   COALESCE(MinimumPayment, 0) as MinimumPayment,
                   COALESCE(InterestRate, 0) as InterestRate,
                   COALESCE(AddedDate, '') as AddedDate,
                   COALESCE(Notes, '') as Notes
            FROM DebtRegister WHERE DebtName = ?
        """, (debt_name,))
        
        if not debt_data:
            flash('Debt not found', 'error')
            return redirect(url_for('debt'))
        
        debt = debt_data[0]
        
        return render_template('edit_debt.html',
                             debt=debt,
                             format_euro=format_euro)
                             
    except Exception as e:
        logger.error(f"Error loading edit debt page: {e}")
        flash(f'Error loading debt: {str(e)}', 'error')
        return redirect(url_for('debt'))

@app.route('/debt/edit/<debt_name>', methods=['POST'])
def update_debt_details(debt_name):
    """Update debt details from edit form"""
    try:
        # Get form data
        current_amount_str = request.form.get('current_amount', '').strip()
        original_amount_str = request.form.get('original_amount', '').strip()
        category = request.form.get('category', '').strip()
        due_date = request.form.get('due_date', '').strip()
        minimum_payment_str = request.form.get('minimum_payment', '').strip()
        interest_rate_str = request.form.get('interest_rate', '').strip()
        notes = request.form.get('notes', '').strip()
        
        # Validation
        errors = []
        
        # Validate amounts
        try:
            current_amount = float(current_amount_str.replace(',', '.'))
            original_amount = float(original_amount_str.replace(',', '.'))
            if current_amount < 0 or original_amount <= 0:
                errors.append("Amounts must be positive")
        except ValueError:
            errors.append("Invalid amount format")
            current_amount = original_amount = None
        
        # Validate optional fields
        minimum_payment = 0
        if minimum_payment_str:
            try:
                minimum_payment = float(minimum_payment_str.replace(',', '.'))
                if minimum_payment < 0:
                    errors.append("Minimum payment cannot be negative")
            except ValueError:
                errors.append("Invalid minimum payment format")
        
        interest_rate = 0
        if interest_rate_str:
            try:
                interest_rate = float(interest_rate_str.replace(',', '.'))
                if interest_rate < 0 or interest_rate > 100:
                    errors.append("Interest rate must be between 0 and 100")
            except ValueError:
                errors.append("Invalid interest rate format")
        
        # Validate due date
        if due_date:
            try:
                datetime.strptime(due_date, '%Y-%m-%d')
            except ValueError:
                errors.append("Invalid due date format")
        
        if errors:
            for error in errors:
                flash(error, 'error')
            return redirect(url_for('edit_debt', debt_name=debt_name))
        
        # Update debt in database
        db_manager.execute_update("""
            UPDATE DebtRegister 
            SET Amount = ?, OriginalDebt = ?, Category = ?, DueDate = ?, 
                MinimumPayment = ?, InterestRate = ?, Notes = ?
            WHERE DebtName = ?
        """, (
            current_amount, original_amount, 
            category if category else None,
            due_date if due_date else None,
            minimum_payment if minimum_payment > 0 else None,
            interest_rate if interest_rate > 0 else None,
            notes if notes else None,
            debt_name
        ), table_name='DebtRegister', record_id=debt_name)
        
        flash('Debt updated successfully', 'success')
        return redirect(url_for('debt'))
        
    except Exception as e:
        logger.error(f"Error updating debt: {e}")
        flash(f'Error updating debt: {str(e)}', 'error')
        return redirect(url_for('edit_debt', debt_name=debt_name))

@app.route('/debt/delete/<debt_name>', methods=['POST'])
def delete_debt(debt_name):
    """Delete a debt completely"""
    try:
        confirmation = request.form.get('confirmation', '').lower()
        
        if confirmation != 'delete':
            flash('Deletion cancelled - you must type "delete" to confirm', 'warning')
            return redirect(url_for('debt'))
        
        # Delete from main register
        db_manager.execute_update(
            "DELETE FROM DebtRegister WHERE DebtName = ?",
            (debt_name,),
            table_name='DebtRegister',
            record_id=debt_name
        )
        
        # Delete the debt database file
        debt_db_path = app.config['DATA_DIR'] / f"{debt_name}.sqlite"
        if debt_db_path.exists():
            debt_db_path.unlink()
        
        flash(f'Debt "{debt_name}" deleted successfully', 'success')
        
    except Exception as e:
        logger.error(f"Error deleting debt: {e}")
        flash(f'Error deleting debt: {str(e)}', 'error')
    
    return redirect(url_for('debt'))

@app.route('/debt/bulk/pay', methods=['POST'])
def bulk_pay_debts():
    """Make payments to multiple debts at once"""
    try:
        debt_names = request.form.getlist('debt_names')
        if not debt_names:
            flash('No debts selected', 'warning')
            return redirect(url_for('debt'))
        
        total_paid = 0
        successful_payments = 0
        
        for debt_name in debt_names:
            payment_amount_str = request.form.get(f'payment_{debt_name}', '').strip()
            if payment_amount_str:
                try:
                    payment_amount = float(payment_amount_str.replace(',', '.'))
                    if payment_amount > 0:
                        # Get current debt amount
                        current_debt = db_manager.execute_query(
                            "SELECT Amount FROM DebtRegister WHERE DebtName = ?", (debt_name,)
                        )
                        
                        if current_debt:
                            current_amount = current_debt[0]['Amount']
                            new_amount = max(0, current_amount - payment_amount)
                            
                            # Update main register
                            db_manager.execute_update(
                                "UPDATE DebtRegister SET Amount = ? WHERE DebtName = ?",
                                (new_amount, debt_name),
                                table_name='DebtRegister',
                                record_id=debt_name
                            )
                            
                            # Add payment record
                            debt_db_path = app.config['DATA_DIR'] / f"{debt_name}.sqlite"
                            if debt_db_path.exists():
                                with sqlite3.connect(debt_db_path) as conn:
                                    conn.execute("""
                                        INSERT INTO DebtSource (Amount, CreatedDate, UNIX, PaymentMethod, Notes) 
                                        VALUES (?, ?, ?, ?, ?)
                                    """, (
                                        payment_amount, 
                                        datetime.now().strftime('%d/%m/%Y'), 
                                        int(datetime.now().timestamp()),
                                        'Bulk Payment',
                                        'Bulk payment operation'
                                    ))
                            
                            total_paid += payment_amount
                            successful_payments += 1
                            
                except ValueError:
                    flash(f'Invalid payment amount for {debt_name}', 'error')
        
        if successful_payments > 0:
            flash(f'Successfully made {successful_payments} payment(s) totaling {format_euro(total_paid)}', 'success')
        else:
            flash('No valid payments were processed', 'warning')
        
    except Exception as e:
        logger.error(f"Error processing bulk payments: {e}")
        flash(f'Error processing bulk payments: {str(e)}', 'error')
    
    return redirect(url_for('debt'))

@app.route('/debt/export', methods=['POST'])
def export_debt_data():
    """Export debt data to CSV"""
    try:
        export_type = request.form.get('export_type', 'summary')
        
        if export_type == 'summary':
            # Export debt summary
            debt_data = db_manager.execute_query("""
                SELECT DebtName, Amount, OriginalDebt, Category, DueDate, 
                       MinimumPayment, InterestRate, Notes 
                FROM DebtRegister ORDER BY DebtName
            """)
            
            output_path = os.path.join(app.config['GENERATED_FOLDER'], f'debt_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')
            
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['DebtName', 'CurrentAmount', 'OriginalDebt', 'TotalPaid', 'Category', 'DueDate', 'MinimumPayment', 'InterestRate', 'Notes']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for debt in debt_data:
                    total_paid = debt['OriginalDebt'] - debt['Amount']
                    writer.writerow({
                        'DebtName': debt['DebtName'],
                        'CurrentAmount': debt['Amount'],
                        'OriginalDebt': debt['OriginalDebt'],
                        'TotalPaid': total_paid,
                        'Category': debt['Category'] or '',
                        'DueDate': debt['DueDate'] or '',
                        'MinimumPayment': debt['MinimumPayment'] or 0,
                        'InterestRate': debt['InterestRate'] or 0,
                        'Notes': debt['Notes'] or ''
                    })
            
            flash(f'Debt summary exported successfully', 'success')
            return send_file(output_path, as_attachment=True, download_name=os.path.basename(output_path))
            
        elif export_type == 'payments':
            # Export payment history for all debts
            output_path = os.path.join(app.config['GENERATED_FOLDER'], f'debt_payments_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')
            
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['DebtName', 'PaymentDate', 'Amount', 'PaymentMethod', 'Notes']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                debt_names = db_manager.execute_query("SELECT DebtName FROM DebtRegister")
                for debt in debt_names:
                    debt_name = debt['DebtName']
                    debt_db_path = app.config['DATA_DIR'] / f"{debt_name}.sqlite"
                    
                    if debt_db_path.exists():
                        with sqlite3.connect(debt_db_path) as conn:
                            conn.row_factory = sqlite3.Row
                            payments = conn.execute("""
                                SELECT CreatedDate, Amount, PaymentMethod, Notes 
                                FROM DebtSource ORDER BY UNIX DESC
                            """).fetchall()
                            
                            for payment in payments:
                                writer.writerow({
                                    'DebtName': debt_name,
                                    'PaymentDate': payment['CreatedDate'],
                                    'Amount': payment['Amount'],
                                    'PaymentMethod': payment['PaymentMethod'] or '',
                                    'Notes': payment['Notes'] or ''
                                })
            
            flash(f'Payment history exported successfully', 'success')
            return send_file(output_path, as_attachment=True, download_name=os.path.basename(output_path))
        
    except Exception as e:
        logger.error(f"Error exporting debt data: {e}")
        flash(f'Error exporting debt data: {str(e)}', 'error')
    
    return redirect(url_for('debt'))

@app.route('/debt/analytics')
def debt_analytics():
    """Show debt analytics and statistics"""
    try:
        # Get debt summary with safe column access
        debt_data = db_manager.execute_query("""
            SELECT DebtName, Amount, OriginalDebt, 
                   COALESCE(Category, 'Uncategorized') as Category,
                   COALESCE(DueDate, '') as DueDate,
                   COALESCE(MinimumPayment, 0) as MinimumPayment,
                   COALESCE(InterestRate, 0) as InterestRate,
                   COALESCE(AddedDate, '') as AddedDate
            FROM DebtRegister ORDER BY Amount DESC
        """)
        
        if not debt_data:
            flash('No debts found', 'info')
            return render_template('debt_analytics.html', analytics={})
        
        # Calculate analytics
        total_current = sum(debt['Amount'] for debt in debt_data)
        total_original = sum(debt['OriginalDebt'] for debt in debt_data)
        total_paid = total_original - total_current
        
        # Progress percentage
        progress_percent = (total_paid / total_original * 100) if total_original > 0 else 0
        
        # Category breakdown
        categories = {}
        for debt in debt_data:
            category = debt['Category'] or 'Uncategorized'
            if category not in categories:
                categories[category] = {'count': 0, 'total': 0, 'original': 0}
            categories[category]['count'] += 1
            categories[category]['total'] += debt['Amount']
            categories[category]['original'] += debt['OriginalDebt']
        
        # Monthly payments needed
        min_monthly = sum(debt['MinimumPayment'] or 0 for debt in debt_data)
        
        # Interest statistics
        weighted_interest = 0
        total_with_interest = 0
        for debt in debt_data:
            if debt['InterestRate'] and debt['InterestRate'] > 0:
                weighted_interest += debt['Amount'] * debt['InterestRate']
                total_with_interest += debt['Amount']
        
        avg_interest = (weighted_interest / total_with_interest) if total_with_interest > 0 else 0
        
        # Time-based analysis
        month_stats = {}
        for debt in debt_data:
            debt_name = debt['DebtName']
            debt_db_path = app.config['DATA_DIR'] / f"{debt_name}.sqlite"
            
            if debt_db_path.exists():
                with sqlite3.connect(debt_db_path) as conn:
                    conn.row_factory = sqlite3.Row
                    payments = conn.execute("""
                        SELECT CreatedDate, Amount FROM DebtSource 
                        WHERE CreatedDate IS NOT NULL
                        ORDER BY UNIX DESC
                    """).fetchall()
                    
                    for payment in payments:
                        try:
                            # Parse date (DD/MM/YYYY format)
                            date_parts = payment['CreatedDate'].split('/')
                            if len(date_parts) == 3:
                                month_key = f"{date_parts[2]}-{date_parts[1].zfill(2)}"
                                if month_key not in month_stats:
                                    month_stats[month_key] = 0
                                month_stats[month_key] += payment['Amount']
                        except:
                            continue
        
        # Sort months
        sorted_months = sorted(month_stats.items(), key=lambda x: x[0], reverse=True)[:12]
        
        analytics = {
            'debt_summary': {
                'total_current': total_current,
                'total_original': total_original,
                'total_paid': total_paid,
                'progress_percent': progress_percent,
                'debt_count': len(debt_data)
            },
            'categories': categories,
            'payments': {
                'min_monthly': min_monthly,
                'avg_interest': avg_interest,
                'total_with_interest': total_with_interest
            },
            'monthly_payments': sorted_months,
            'debt_list': debt_data
        }
        
        return render_template('debt_analytics.html', analytics=analytics)
        
    except Exception as e:
        logger.error(f"Error generating debt analytics: {e}")
        flash(f'Error generating analytics: {str(e)}', 'error')
        return redirect(url_for('debt'))

@app.route('/debt/report')
def debt_report():
    """Generate comprehensive debt report"""
    try:
        report_data = {
            'generated_date': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'debts': [],
            'totals': {'current': 0, 'original': 0, 'paid': 0},
            'categories': {},
            'payment_trends': []
        }
        
        # Get all debts with safe column access
        debt_data = db_manager.execute_query("""
            SELECT DebtName, Amount, OriginalDebt, 
                   COALESCE(Category, 'Uncategorized') as Category,
                   COALESCE(DueDate, '') as DueDate,
                   COALESCE(MinimumPayment, 0) as MinimumPayment,
                   COALESCE(InterestRate, 0) as InterestRate,
                   COALESCE(AddedDate, '') as AddedDate,
                   COALESCE(Notes, '') as Notes
            FROM DebtRegister ORDER BY Amount DESC
        """)
        
        for debt in debt_data:
            debt_name = debt['DebtName']
            debt_info = {
                'name': debt_name,
                'current_amount': debt['Amount'],
                'original_amount': debt['OriginalDebt'],
                'paid_amount': debt['OriginalDebt'] - debt['Amount'],
                'category': debt['Category'] or 'Uncategorized',
                'due_date': debt['DueDate'],
                'minimum_payment': debt['MinimumPayment'] or 0,
                'interest_rate': debt['InterestRate'] or 0,
                'notes': debt['Notes'],
                'payments': []
            }
            
            # Get payment history
            debt_db_path = app.config['DATA_DIR'] / f"{debt_name}.sqlite"
            if debt_db_path.exists():
                with sqlite3.connect(debt_db_path) as conn:
                    conn.row_factory = sqlite3.Row
                    payments = conn.execute("""
                        SELECT CreatedDate, Amount, PaymentMethod, Notes 
                        FROM DebtSource ORDER BY UNIX DESC LIMIT 5
                    """).fetchall()
                    
                    debt_info['payments'] = [
                        {
                            'date': payment['CreatedDate'],
                            'amount': payment['Amount'],
                            'method': payment['PaymentMethod'],
                            'notes': payment['Notes']
                        } for payment in payments
                    ]
            
            report_data['debts'].append(debt_info)
            
            # Update totals
            report_data['totals']['current'] += debt['Amount']
            report_data['totals']['original'] += debt['OriginalDebt']
            report_data['totals']['paid'] += debt['OriginalDebt'] - debt['Amount']
            
            # Update categories
            category = debt['Category'] or 'Uncategorized'
            if category not in report_data['categories']:
                report_data['categories'][category] = {'count': 0, 'amount': 0}
            report_data['categories'][category]['count'] += 1
            report_data['categories'][category]['amount'] += debt['Amount']
        
        return render_template('debt_report.html', report=report_data)
        
    except Exception as e:
        logger.error(f"Error generating debt report: {e}")
        flash(f'Error generating report: {str(e)}', 'error')
        return redirect(url_for('debt'))

@app.route('/debt/log/<debt_name>')
def debt_log(debt_name):
    """View debt payment log"""
    debt_db_path = app.config['DATA_DIR'] / f"{debt_name}.sqlite"
    
    if not debt_db_path.exists():
        flash('Debt log not found', 'error')
        return redirect(url_for('debt'))
    
    with sqlite3.connect(debt_db_path) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.execute("SELECT CreatedDate, Amount FROM DebtSource ORDER BY UNIX DESC")
        log_data = cursor.fetchall()
    
    return render_template('debt_log.html',
                         debt_name=debt_name,
                         log_data=log_data,
                         format_euro=format_euro)

@app.route('/api/invoices/import', methods=['POST'])
def import_invoices():
    """Import invoices from Excel or ZIP files (supports multiple files via multiple requests)"""
    try:
        logger.info("Import request received")
        
        if 'file' not in request.files:
            logger.error("No file in request")
            return jsonify({'status': 'error', 'message': 'No file selected', 'silent': True})
        
        file = request.files['file']
        if file.filename == '':
            logger.error("Empty filename")
            return jsonify({'status': 'error', 'message': 'No file selected', 'silent': True})
        
        logger.info(f"Processing file: {file.filename}")
        
        # Get file extension
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        
        logger.info(f"File extension: {file_ext}")
        
        if file_ext not in ['xlsx', 'xls', 'zip']:
            return jsonify({'status': 'error', 'message': f'File "{filename}": Only Excel (.xlsx, .xls) and ZIP files are allowed', 'silent': True})
        
        # Save uploaded file temporarily with unique name to avoid conflicts
        upload_folder = app.config['UPLOAD_FOLDER']
        os.makedirs(upload_folder, exist_ok=True)  # Ensure folder exists
        
        # Create unique filename to prevent concurrent upload conflicts
        import time
        import uuid
        timestamp = str(int(time.time() * 1000000))  # Microsecond precision
        unique_id = str(uuid.uuid4())[:8]  # Short unique ID
        name_part = filename.rsplit('.', 1)[0] if '.' in filename else filename
        ext_part = f".{file_ext}" if file_ext else ""
        unique_filename = f"{name_part}_{timestamp}_{unique_id}{ext_part}"
        
        temp_path = os.path.join(upload_folder, unique_filename)
        logger.info(f"Saving to temp path: {temp_path} (original: {filename})")
        file.save(temp_path)
        
        imported_count = 0
        
        if file_ext in ['xlsx', 'xls']:
            # Process Excel file
            logger.info("Processing Excel file")
            imported_count = process_excel_import(temp_path)
            logger.info(f"Imported count: {imported_count}")
        elif file_ext == 'zip':
            # Process ZIP file containing Excel files
            logger.info("Processing ZIP file")
            imported_count = process_zip_import(temp_path)
            logger.info(f"Imported count from ZIP: {imported_count}")
        
        # Clean up temporary file
        if os.path.exists(temp_path):
            os.remove(temp_path)
            logger.info("Cleaned up temporary file")
        
        if imported_count > 0:
            return jsonify({
                'status': 'success', 
                'message': f'Successfully imported {imported_count} invoice(s) from "{filename}"',
                'silent': True,  # Flag to prevent popup
                'imported_count': imported_count,
                'filename': filename
            })
        else:
            return jsonify({
                'status': 'error', 
                'message': f'No valid invoice data found in "{filename}"',
                'silent': True,  # Flag to prevent popup
                'filename': filename
            })
            
    except Exception as e:
        logger.error(f"Error importing invoices: {e}")
        return jsonify({'status': 'error', 'message': f'Import failed for "{filename}": {str(e)}', 'silent': True})

def process_excel_import(file_path):
    """Process Excel file and import invoice data"""
    try:
        logger.info(f"Loading Excel file: {file_path}")
        
        # First, validate that the file exists and has content
        if not os.path.exists(file_path):
            logger.error(f"File does not exist: {file_path}")
            return 0
            
        file_size = os.path.getsize(file_path)
        logger.info(f"File size: {file_size} bytes")
        
        if file_size == 0:
            logger.error("File is empty")
            return 0
            
        # Try to detect file type by reading first few bytes
        try:
            with open(file_path, 'rb') as f:
                file_header = f.read(8)
                logger.info(f"File header (hex): {file_header.hex()}")
                
                # Check if it's a ZIP file (Excel .xlsx files start with PK)
                if not file_header.startswith(b'PK'):
                    logger.error("File does not appear to be a valid Excel .xlsx file (missing ZIP signature)")
                    # Try to read as text to see what it contains
                    try:
                        with open(file_path, 'r', encoding='utf-8', errors='ignore') as text_file:
                            content_preview = text_file.read(200)
                            logger.info(f"File content preview: {content_preview}")
                    except:
                        pass
                    return 0
        except Exception as e:
            logger.error(f"Error reading file header: {e}")
            return 0
        
        # Load workbook with data_only=True to get calculated values instead of formulas
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except zipfile.BadZipFile as e:
            logger.error(f"Invalid Excel file format - file appears to be corrupted: {e}")
            return 0
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return 0
            
        worksheet = workbook.active
        
        imported_count = 0
        
        # Debug: Print sheet info
        logger.info(f"Worksheet name: {worksheet.title}")
        logger.info(f"Max row: {worksheet.max_row}, Max column: {worksheet.max_column}")
        
        # Method 1: Try Template.xlsx format (C13, C14, F18)
        logger.info("=== Trying Template.xlsx format ===")
        invoice_number = worksheet['C13'].value
        invoice_date = worksheet['C14'].value  
        amount_excl = worksheet['F18'].value
        
        logger.info(f"C13 (Invoice Number): {invoice_number} (type: {type(invoice_number)})")
        logger.info(f"C14 (Invoice Date): {invoice_date} (type: {type(invoice_date)})")
        logger.info(f"F18 (Amount): {amount_excl} (type: {type(amount_excl)})")
        
        # Also check alternative cells that might contain data
        logger.info("=== Checking alternative cells ===")
        f43_value = worksheet['F43'].value
        f44_value = worksheet['F44'].value
        f45_value = worksheet['F45'].value
        logger.info(f"F43: {f43_value} (type: {type(f43_value)})")
        logger.info(f"F44: {f44_value} (type: {type(f44_value)})")
        logger.info(f"F45: {f45_value} (type: {type(f45_value)})")
        
        def parse_european_number(value):
            """Convert European number format (comma as decimal) to float"""
            if value is None:
                return None
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, str):
                # Handle European format: 666,75 -> 666.75
                # Remove any spaces and replace comma with dot
                cleaned = str(value).strip().replace(' ', '').replace(',', '.')
                try:
                    return float(cleaned)
                except ValueError:
                    return None
            return None
        
        # Try to parse the amount with European number format
        parsed_amount = parse_european_number(amount_excl)
        logger.info(f"Parsed amount: {parsed_amount}")
        
        if invoice_number and parsed_amount is not None and str(invoice_number).strip():
            try:
                # Validate and convert data
                amount_excl_final = float(parsed_amount)
                btw = round(amount_excl_final * 0.21, 2)  # 21% BTW
                incl = round(amount_excl_final + btw, 2)
                
                # Convert invoice number to integer to remove any decimal places (e.g., 240007.0 -> 240007)
                invoice_number_clean = str(int(float(invoice_number)))
                
                # Convert date format if needed
                if isinstance(invoice_date, datetime):
                    formatted_date = invoice_date.strftime('%d-%m-%Y')
                elif isinstance(invoice_date, str):
                    # Try to parse string date
                    try:
                        parsed_date = datetime.strptime(str(invoice_date), '%d-%m-%Y')
                        formatted_date = parsed_date.strftime('%d-%m-%Y')
                    except:
                        try:
                            parsed_date = datetime.strptime(str(invoice_date), '%Y-%m-%d')
                            formatted_date = parsed_date.strftime('%d-%m-%Y')
                        except:
                            formatted_date = datetime.now().strftime('%d-%m-%Y')
                else:
                    formatted_date = datetime.now().strftime('%d-%m-%Y')
                
                logger.info(f"Final processed data - Number: {invoice_number_clean}, Date: {formatted_date}, Amount: {amount_excl_final}")
                
                # Check if invoice already exists
                existing = db_manager.execute_query(
                    "SELECT COUNT(*) as count FROM Invoices WHERE InvoiceID = ?", 
                    (invoice_number_clean,)
                )
                
                if existing[0]['count'] == 0:
                    # Insert new invoice
                    db_manager.execute_update(
                        "INSERT INTO Invoices (InvoiceID, InvoiceDate, Excl, BTW, Incl) VALUES (?, ?, ?, ?, ?)",
                        (invoice_number_clean, formatted_date, amount_excl_final, btw, incl)
                    )
                    imported_count += 1
                    logger.info(f"Successfully imported invoice {invoice_number_clean}: €{amount_excl_final} (Excl)")
                else:
                    logger.info(f"Invoice {invoice_number_clean} already exists, skipping")
                    
            except (ValueError, TypeError) as e:
                logger.error(f"Error processing invoice data: {e}")
        
        # If F18 didn't work, try F43 (total excl BTW)
        if not parsed_amount:
            parsed_f43 = parse_european_number(f43_value)
            logger.info(f"Trying F43 amount: {parsed_f43}")
            
            if invoice_number and parsed_f43 is not None and str(invoice_number).strip():
                try:
                    amount_excl_final = float(parsed_f43)
                    btw = round(amount_excl_final * 0.21, 2)
                    incl = round(amount_excl_final + btw, 2)
                    
                    # Convert invoice number to integer to remove any decimal places
                    invoice_number_clean = str(int(float(invoice_number)))
                    
                    formatted_date = str(invoice_date) if invoice_date else datetime.now().strftime('%d-%m-%Y')
                    
                    existing = db_manager.execute_query(
                        "SELECT COUNT(*) as count FROM Invoices WHERE InvoiceID = ?", 
                        (invoice_number_clean,)
                    )
                    
                    if existing[0]['count'] == 0:
                        db_manager.execute_update(
                            "INSERT INTO Invoices (InvoiceID, InvoiceDate, Excl, BTW, Incl) VALUES (?, ?, ?, ?, ?)",
                            (invoice_number_clean, formatted_date, amount_excl_final, btw, incl)
                        )
                        imported_count += 1
                        logger.info(f"Successfully imported invoice {invoice_number_clean} from F43: €{amount_excl_final} (Excl)")
                    else:
                        logger.info(f"Invoice {invoice_number_clean} already exists")
                        
                except (ValueError, TypeError) as e:
                    logger.error(f"Error processing F43 data: {e}")
        
        # Method 2: Try scanning all cells for data (only if no invoices imported yet)
        if imported_count == 0:
            logger.info("=== Template format not found, scanning all cells ===")
            
            # Look for any numeric values that could be invoice numbers or amounts
            potential_invoices = []
            potential_amounts = []
            
            for row in range(1, min(51, worksheet.max_row + 1)):
                for col in range(1, min(26, worksheet.max_column + 1)):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None:
                        # Check for potential invoice numbers (integers with 4+ digits)
                        if isinstance(cell.value, (int, float)) and str(int(cell.value)).isdigit() and len(str(int(cell.value))) >= 4:
                            potential_invoices.append((cell.coordinate, int(cell.value)))
                        
                        # Check for potential amounts (try to parse European format)
                        parsed_val = parse_european_number(cell.value)
                        if parsed_val is not None and 1 <= parsed_val <= 100000:
                            potential_amounts.append((cell.coordinate, parsed_val))
            
            logger.info(f"Found {len(potential_invoices)} potential invoice numbers: {potential_invoices[:5]}")
            logger.info(f"Found {len(potential_amounts)} potential amounts: {potential_amounts[:5]}")
            
            # Try to match invoice numbers with amounts
            if potential_invoices and potential_amounts:
                # Use the first potential combination
                invoice_num = potential_invoices[0][1]
                amount = potential_amounts[0][1]
                
                logger.info(f"Attempting to import: Invoice {invoice_num}, Amount {amount}")
                
                # Check if invoice already exists
                existing = db_manager.execute_query(
                    "SELECT COUNT(*) as count FROM Invoices WHERE InvoiceID = ?", 
                    (str(int(invoice_num)),)
                )
                
                if existing[0]['count'] == 0:
                    btw = round(amount * 0.21, 2)
                    incl = round(amount + btw, 2)
                    current_date = datetime.now().strftime('%d-%m-%Y')
                    
                    db_manager.execute_update(
                        "INSERT INTO Invoices (InvoiceID, InvoiceDate, Excl, BTW, Incl) VALUES (?, ?, ?, ?, ?)",
                        (str(int(invoice_num)), current_date, amount, btw, incl)
                    )
                    imported_count += 1
                    logger.info(f"Successfully imported scanned invoice {int(invoice_num)}: €{amount}")
                else:
                    logger.info(f"Invoice {int(invoice_num)} already exists")
        else:
            logger.info("=== Skipping cell scanning - invoice already imported from template format ===")
        
        return imported_count
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0

def process_zip_import(zip_path):
    """Process ZIP file containing Excel files and import invoice data"""
    try:
        logger.info(f"Processing ZIP file: {zip_path}")
        
        # Validate ZIP file exists and has content
        if not os.path.exists(zip_path):
            logger.error(f"ZIP file does not exist: {zip_path}")
            return 0
            
        file_size = os.path.getsize(zip_path)
        logger.info(f"ZIP file size: {file_size} bytes")
        
        if file_size == 0:
            logger.error("ZIP file is empty")
            return 0
        
        total_imported = 0
        
        # Create temporary directory for extracting files
        with tempfile.TemporaryDirectory() as temp_dir:
            logger.info(f"Created temporary directory: {temp_dir}")
            
            # Extract ZIP file
            try:
                with zipfile.ZipFile(zip_path, 'r') as zip_file:
                    # Get list of files in ZIP
                    file_list = zip_file.namelist()
                    logger.info(f"Files in ZIP: {file_list}")
                    
                    # Filter for Excel files (including those in subdirectories)
                    excel_files = []
                    for f in file_list:
                        # Skip macOS metadata files and directories
                        if f.startswith('__MACOSX/') or f.endswith('/'):
                            continue
                        # Check if it's an Excel file
                        if f.lower().endswith(('.xlsx', '.xls')):
                            excel_files.append(f)
                    
                    logger.info(f"Found {len(excel_files)} Excel files: {excel_files}")
                    
                    if not excel_files:
                        logger.warning("No Excel files found in ZIP archive")
                        return 0
                    
                    # Extract and process each Excel file
                    for excel_file in excel_files:
                        try:
                            logger.info(f"Processing Excel file from ZIP: {excel_file}")
                            
                            # Extract the file
                            extracted_path = zip_file.extract(excel_file, temp_dir)
                            logger.info(f"Extracted to: {extracted_path}")
                            
                            # Process the Excel file
                            imported_count = process_excel_import(extracted_path)
                            total_imported += imported_count
                            
                            logger.info(f"Imported {imported_count} invoices from {excel_file}")
                            
                        except Exception as e:
                            logger.error(f"Error processing Excel file {excel_file} from ZIP: {e}")
                            continue  # Continue with next file
                            
            except zipfile.BadZipFile as e:
                logger.error(f"Invalid ZIP file format: {e}")
                return 0
            except Exception as e:
                logger.error(f"Error extracting ZIP file: {e}")
                return 0
        
        logger.info(f"Total imported from ZIP: {total_imported} invoices")
        return total_imported
        
    except Exception as e:
        logger.error(f"Error processing ZIP file: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0

@app.route('/settings')
def settings():
    """Application settings and configuration page"""
    try:
        # Get application statistics for settings overview
        app_stats = {
            'database_size': get_database_size(),
            'total_invoices': db_manager.execute_query("SELECT COUNT(*) as count FROM Invoices")[0]['count'],
            'total_withdraws': db_manager.execute_query("SELECT COUNT(*) as count FROM Withdraw")[0]['count'],
            'total_debts': db_manager.execute_query("SELECT COUNT(*) as count FROM DebtRegister")[0]['count'],
            'total_quarters': db_manager.execute_query("SELECT COUNT(*) as count FROM KwartaalData")[0]['count'],
        }
        
        # Get configuration info
        config_info = {
            'upload_folder': app.config['UPLOAD_FOLDER'],
            'generated_folder': app.config['GENERATED_FOLDER'],
            'max_file_size': app.config['MAX_CONTENT_LENGTH'] // (1024 * 1024),  # Convert to MB
            'database_path': getattr(app.config, 'DATABASE_PATH', 'Not configured'),
            'debug_mode': app.config.get('DEBUG', False),
            'environment': config_name
        }
        
        return render_template('settings.html',
                             app_stats=app_stats,
                             config_info=config_info,
                             format_euro=format_euro)
                             
    except Exception as e:
        logger.error(f"Error loading settings: {e}")
        flash(f'Error loading settings: {str(e)}', 'error')
        return render_template('settings.html',
                             app_stats={},
                             config_info={},
                             format_euro=format_euro)

def get_database_size():
    """Get the size of the database file in MB"""
    try:
        db_path = app.config.get('DATABASE_PATH', 'ozark_finances.db')
        if os.path.exists(db_path):
            size_bytes = os.path.getsize(db_path)
            size_mb = size_bytes / (1024 * 1024)
            return round(size_mb, 2)
        return 0
    except Exception:
        return 0

@app.route('/settings/export-data', methods=['POST'])
def export_data():
    """Export application data to CSV"""
    try:
        export_type = request.form.get('export_type')
        
        if export_type == 'invoices':
            # Export invoices data
            invoices = db_manager.execute_query("SELECT * FROM Invoices ORDER BY CAST(InvoiceID AS INTEGER) DESC")
            
            output_path = os.path.join(app.config['GENERATED_FOLDER'], f'invoices_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')
            
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                if invoices:
                    fieldnames = invoices[0].keys()
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    for invoice in invoices:
                        writer.writerow(dict(invoice))
            
            flash(f'Invoices exported successfully to {os.path.basename(output_path)}', 'success')
            return send_file(output_path, as_attachment=True, download_name=os.path.basename(output_path))
            
        elif export_type == 'withdraws':
            # Export withdraws data
            withdraws = db_manager.execute_query("SELECT * FROM Withdraw ORDER BY Date DESC")
            
            output_path = os.path.join(app.config['GENERATED_FOLDER'], f'withdraws_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')
            
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                if withdraws:
                    fieldnames = withdraws[0].keys()
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    for withdraw in withdraws:
                        writer.writerow(dict(withdraw))
            
            flash(f'Withdraws exported successfully to {os.path.basename(output_path)}', 'success')
            return send_file(output_path, as_attachment=True, download_name=os.path.basename(output_path))
            
        elif export_type == 'all':
            # Export all data to a zip file
            import zipfile
            
            zip_path = os.path.join(app.config['GENERATED_FOLDER'], f'complete_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip')
            
            with zipfile.ZipFile(zip_path, 'w') as zipf:
                # Export each table
                tables = ['Invoices', 'Withdraw', 'DebtRegister', 'KwartaalData']
                
                for table in tables:
                    try:
                        data = db_manager.execute_query(f"SELECT * FROM {table}")
                        if data:
                            csv_content = []
                            fieldnames = data[0].keys()
                            
                            # Create CSV content in memory
                            import io
                            output = io.StringIO()
                            writer = csv.DictWriter(output, fieldnames=fieldnames)
                            writer.writeheader()
                            for row in data:
                                writer.writerow(dict(row))
                            
                            # Add to zip
                            zipf.writestr(f'{table.lower()}.csv', output.getvalue())
                    except Exception as e:
                        logger.error(f"Error exporting table {table}: {e}")
            
            flash(f'Complete data export created successfully', 'success')
            return send_file(zip_path, as_attachment=True, download_name=os.path.basename(zip_path))
        
        else:
            flash('Invalid export type selected', 'error')
            
    except Exception as e:
        logger.error(f"Error exporting data: {e}")
        flash(f'Error exporting data: {str(e)}', 'error')
    
    return redirect(url_for('settings'))

@app.route('/settings/clear-data', methods=['POST'])
def clear_data():
    """Clear specific data tables with confirmation"""
    try:
        clear_type = request.form.get('clear_type')
        confirmation = request.form.get('confirmation', '').lower()
        
        if confirmation != 'delete':
            flash('Please type "delete" to confirm data deletion', 'error')
            return redirect(url_for('settings'))
        
        if clear_type == 'generated_files':
            # Clear generated folder
            generated_files = glob.glob(os.path.join(app.config['GENERATED_FOLDER'], '*'))
            deleted_count = 0
            
            for file_path in generated_files:
                try:
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                        deleted_count += 1
                except Exception as e:
                    logger.error(f"Error deleting file {file_path}: {e}")
            
            flash(f'Deleted {deleted_count} generated files', 'success')
            
        elif clear_type == 'uploads':
            # Clear uploads folder
            upload_files = glob.glob(os.path.join(app.config['UPLOAD_FOLDER'], '*'))
            deleted_count = 0
            
            for file_path in upload_files:
                try:
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                        deleted_count += 1
                except Exception as e:
                    logger.error(f"Error deleting file {file_path}: {e}")
            
            flash(f'Deleted {deleted_count} uploaded files', 'success')
            
        elif clear_type == 'imported_data':
            # Clear imported invoice data from database
            try:
                # Get count before deletion
                count_result = db_manager.execute_query("SELECT COUNT(*) as count FROM Invoices")
                total_invoices = count_result[0]['count'] if count_result else 0
                
                # Clear all invoices
                db_manager.execute_update("DELETE FROM Invoices")
                flash(f'Deleted {total_invoices} imported invoice records from database', 'success')
            except Exception as e:
                logger.error(f"Error clearing invoice data: {e}")
                flash(f'Error clearing invoice data: {str(e)}', 'error')
                
        elif clear_type == 'withdraw_data':
            # Clear all withdraw data from database
            try:
                # Get count before deletion
                count_result = db_manager.execute_query("SELECT COUNT(*) as count FROM Withdraw")
                total_withdraws = count_result[0]['count'] if count_result else 0
                
                # Clear all withdraws
                db_manager.execute_update("DELETE FROM Withdraw")
                flash(f'Deleted {total_withdraws} withdraw records from database', 'success')
            except Exception as e:
                logger.error(f"Error clearing withdraw data: {e}")
                flash(f'Error clearing withdraw data: {str(e)}', 'error')
            
        else:
            flash('Invalid clear type selected', 'error')
            
    except Exception as e:
        logger.error(f"Error clearing data: {e}")
        flash(f'Error clearing data: {str(e)}', 'error')
    
    return redirect(url_for('settings'))

@app.route('/generate-invoice')
def generate_invoice():
    """Invoice generator page"""
    # Get next invoice number
    next_invoice_number = get_next_invoice_number()
    current_date = date.today().strftime('%Y-%m-%d')  # HTML date input format
    
    return render_template('generate_invoice.html',
                         next_invoice_number=next_invoice_number,
                         current_date=current_date,
                         format_euro=format_euro)

@app.route('/generate-invoice', methods=['POST'])
def create_invoice():
    """Handle invoice creation (simplified FlaskWebApp-style)"""
    try:
        # Get output filename or generate auto filename
        output_filename = request.form.get('output_filename', '').strip()
        invoice_number = get_next_invoice_number()
        current_date = datetime.now().strftime('%d-%m-%Y')
        
        if not output_filename:
            # Auto-generate filename like FlaskWebApp
            output_filename = f"TringTring_{invoice_number}_{current_date}_TijnBakker"
        
        if not output_filename.endswith('.xlsx'):
            output_filename += '.xlsx'
        
        # Get F18 amount (only required field like FlaskWebApp)
        amount_excl_str = request.form.get('cell_f18', '0').strip()
        
        # Validate F18 amount (main requirement) with European format support
        errors = []
        
        if not amount_excl_str:
            errors.append('F18 amount is required')
        else:
            try:
                # Handle European decimal format (comma as decimal separator)
                cleaned_amount = amount_excl_str.replace(',', '.')
                amount_excl = float(cleaned_amount)
                if amount_excl <= 0:
                    errors.append('F18 amount must be greater than 0')
            except ValueError:
                errors.append('Invalid amount format in F18 (use comma or dot for decimals)')
                amount_excl = 0
        
        if errors:
            for error in errors:
                flash(error, 'error')
            return redirect(url_for('generate_invoice'))
        
        # Simple cell data (matching FlaskWebApp minimal approach)
        cell_data = {
            'invoice_number': invoice_number,
            'invoice_date': current_date,
            'filename': output_filename,
            'amount_excl': amount_excl_str,  # F18: Main amount (only required field)
        }
        
        # Handle optional image upload
        image_path = None
        if 'invoice_image' in request.files:
            file = request.files['invoice_image']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(image_path)
        
        # Create Excel file using FlaskWebApp method
        excel_path = create_excel_from_template(cell_data, image_path)
        
        if excel_path:
            # Calculate database values (like FlaskWebApp)
            excl = round(amount_excl, 2)
            btw = round(amount_excl * 0.21, 2)  # 21% BTW
            incl = round(excl + btw, 2)
            
            # Save to database (convert date format for database)
            db_manager.execute_update(
                "INSERT INTO Invoices (InvoiceID, InvoiceDate, Excl, BTW, Incl) VALUES (?, ?, ?, ?, ?)",
                (invoice_number, current_date, excl, btw, incl),
                table_name='Invoices',
                record_id=invoice_number
            )
            
            flash(f'Invoice {invoice_number} generated successfully! Amount: €{excl:.2f} (Excl), €{incl:.2f} (Incl)', 'success')
            
            # Return the Excel file for download
            return send_file(excel_path, as_attachment=True, download_name=output_filename)
        else:
            flash('Error generating Excel file. Please check your data and try again.', 'error')
            return redirect(url_for('generate_invoice'))
            
    except Exception as e:
        logger.error(f"Error creating invoice: {e}")
        flash(f'Error creating invoice: {str(e)}', 'error')
        return redirect(url_for('generate_invoice'))

def process_knab_csv_data(input_file_path):
    """
    Process Knab CSV file and extract only Transactiedatum, Bedrag, and Omschrijving fields.
    
    Args:
        input_file_path (str): Path to the input Knab CSV file
    
    Returns:
        dict: Result dictionary with success status, output path, and statistics
    """
    import re
    from io import StringIO
    
    try:
        # Generate output filename
        base_name = os.path.splitext(os.path.basename(input_file_path))[0]
        output_filename = f"{base_name}_cleaned.csv"
        output_path = os.path.join(app.config['GENERATED_FOLDER'], output_filename)
        
        # Track statistics
        total_rows = 0
        processed_rows = 0
        skipped_rows = 0
        
        with open(input_file_path, 'r', encoding='utf-8') as infile:
            content = infile.read()
            
        # Split into lines
        lines = content.strip().split('\n')
        
        # Find the header line (should contain "Rekeningnummer;Transactiedatum;...")
        header_line_index = -1
        for i, line in enumerate(lines):
            if 'Rekeningnummer' in line and 'Transactiedatum' in line and 'Omschrijving' in line:
                header_line_index = i
                break
        
        if header_line_index == -1:
            return {
                'success': False,
                'error': 'Could not find valid Knab CSV header. Please ensure this is a Knab bank export file.'
            }
        
        # Parse header to find column positions
        header_line = lines[header_line_index]
        headers = [h.strip('"') for h in header_line.split(';')]
        
        # Find the indices of our target columns
        try:
            transactiedatum_idx = headers.index('Transactiedatum')
            bedrag_idx = headers.index('Bedrag')
            omschrijving_idx = headers.index('Omschrijving')
        except ValueError as e:
            return {
                'success': False,
                'error': f'Required column not found in CSV file: {e}'
            }
        
        # Process data lines
        processed_data = []
        
        # Start from the line after the header
        for line_num, line in enumerate(lines[header_line_index + 1:], start=header_line_index + 2):
            total_rows += 1
            
            # Skip empty lines
            if not line.strip():
                skipped_rows += 1
                continue
            
            # Parse the line using CSV reader for proper quote handling
            try:
                csv_reader = csv.reader(StringIO(line), delimiter=';')
                row = next(csv_reader)
                
                # Check if we have enough columns
                if len(row) <= max(transactiedatum_idx, bedrag_idx, omschrijving_idx):
                    skipped_rows += 1
                    continue
                
                # Extract the target fields
                transactiedatum = row[transactiedatum_idx].strip('"').strip()
                bedrag = row[bedrag_idx].strip('"').strip()
                omschrijving = row[omschrijving_idx].strip('"').strip()
                
                # Validate date format (DD-MM-YYYY)
                if not re.match(r'\d{2}-\d{2}-\d{4}', transactiedatum):
                    skipped_rows += 1
                    continue
                
                # Validate amount (should be numeric, possibly with comma as decimal separator)
                if not re.match(r'^\d+([,\.]\d+)?$', bedrag):
                    skipped_rows += 1
                    continue
                
                # Add to processed data
                processed_data.append([transactiedatum, bedrag, omschrijving])
                processed_rows += 1
                
            except Exception:
                skipped_rows += 1
                continue
        
        # Write the processed data to output file
        with open(output_path, 'w', newline='', encoding='utf-8') as outfile:
            writer = csv.writer(outfile, delimiter=',')  # Use comma delimiter for output
            
            # Write header
            writer.writerow(['Transactiedatum', 'Bedrag', 'Omschrijving'])
            
            # Write data
            for row in processed_data:
                writer.writerow(row)
        
        return {
            'success': True,
            'output_path': output_path,
            'output_filename': output_filename,
            'statistics': {
                'total_rows': total_rows,
                'processed_rows': processed_rows,
                'skipped_rows': skipped_rows
            }
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@app.context_processor
def utility_processor():
    """Make utility functions available in templates"""
    return dict(
        format_euro=format_euro,
        today=datetime.now().strftime('%Y-%m-%d')
    )

if __name__ == '__main__':
    app.run(debug=app.config['DEBUG'], host=app.config['HOST'], port=app.config['PORT'])
